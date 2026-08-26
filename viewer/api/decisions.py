"""Reading a spreadsheet of decisions back in.

The export goes out as a CSV that somebody opens in Excel, fills in and sends
back. This is the half that reads it: it turns whatever arrives — the CSV
untouched, or the same thing saved as .xlsx — into a header row and numbered
rows of text, and refuses clearly when it is neither.

It is a separate module from main.py because it is the only code here that
parses a file format. Nothing in it knows what a verdict is or what closing one
means; it hands back cells, and main.py decides. Keeping the two apart is what
lets the vocabulary and the schema semantics stay next to the table they belong
to rather than next to a ZIP sniffer.

**Nothing here trusts the filename.** The loop this exists for is export, open
in Excel, save, upload, and Excel is perfectly willing to write a CSV to a name
ending .xlsx and the other way round. What the file *is* is decided by its first
bytes, which cannot be renamed.
"""

from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

# A file this size is not a month of a review, it is a mistake — a database
# dump, a video, a directory somebody zipped. The cap is here rather than at the
# route because it is a fact about what this parser is for, and because both
# readers below would otherwise pull the whole thing into memory before finding
# out.
MAX_BYTES = 8 * 1024 * 1024

# The first four bytes of every ZIP, and therefore of every .xlsx: the format is
# a zipped directory of XML.
_ZIP = b"PK\x03\x04"

# The OLE2 compound-document header, which is the *old* .xls. openpyxl does not
# read it and never will, and that is worth telling apart from a corrupt .xlsx
# because the answer is different: save it again as .xlsx or as CSV.
_OLE2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class FileError(ValueError):
    """The upload is not a sheet this can read, and why."""


@dataclass
class Sheet:
    """A header row and the rows under it, as text, with Excel's own numbering.

    `number` is the row as the person staring at Excel sees it: the header is
    row 1 and the first row of data is row 2. Every message about a row quotes
    that number back at them, so it is carried out of the parser rather than
    recomputed by whoever formats the message — which is how one place ends up
    off by one and the other does not.
    """

    kind: str                       # "csv" or "xlsx"
    header: List[str]               # folded, in file order
    rows: List[Tuple[int, List[str]]] = field(default_factory=list)
    # Which tab of the workbook this came from. Empty for a CSV, which has only
    # one and so has nothing to disambiguate.
    tab: str = ""


def digest(payload: bytes) -> str:
    """A name for these exact bytes, which the preview and the apply must agree on.

    Not a security measure and not trying to be. It is what lets an apply prove
    it is carrying the file that was previewed rather than the one next to it in
    the downloads folder.
    """
    return hashlib.sha256(payload).hexdigest()


def fold(text) -> str:
    """The form two spellings of one word agree on.

    Used for column names and for decision values alike, and that is the whole
    trick that makes the round trip work: the export writes the *label* — "Our
    verdict", "Approved — no action" — while the store holds the key,
    `approved-no-action`. Folding both to lowercase words joined by single
    hyphens lands them on the same string without this module holding a copy of
    the label table, or a literal em dash, or an opinion about how many spaces
    somebody left around it.
    """
    return re.sub(r"[^a-z0-9]+", "-", str(text or "").lower()).strip("-")


def _text(value) -> str:
    """One cell as the text it would have been in the CSV.

    openpyxl hands back what the cell *is*, which for an id column is an int and
    for a date is a datetime. A float that is a whole number is rendered without
    the .0, because 812.0 is what Excel makes of an id it decided was a number,
    and "no verdict with id 812.0" is a message that helps nobody.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _delimiter(line: str) -> str:
    """What this CSV separates fields with.

    The export writes commas. Excel does not always: on a machine whose locale
    uses the comma as a decimal separator, "Save as CSV" writes semicolons, and
    what comes back is the same file with a different punctuation mark. Counting
    on the header line settles it — no column name in this export contains any
    of the three — and the alternative is not a subtle failure but a file whose
    every row is one field and whose id column has therefore vanished.
    """
    counts = {sep: line.count(sep) for sep in (",", ";", "\t")}
    best = max(counts, key=lambda sep: counts[sep])
    return best if counts[best] else ","


def _from_csv(payload: bytes) -> Sheet:
    try:
        # utf-8-sig, because the export puts a byte order mark in front for
        # Excel's sake, and utf-8 alone would leave it glued to the front of the
        # first column name — where it turns "id" into something matching nothing.
        text = payload.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise FileError(
            "That file is neither a spreadsheet nor text this can read. Upload "
            "the CSV the export gave you, or that file saved out of Excel as "
            ".xlsx or .csv.")
    if "\x00" in text[:4096]:
        # UTF-16 and most binaries decode without raising and arrive here
        # looking like text, then parse into nonsense. Said plainly, because
        # "no id column" would be a true and completely unhelpful answer to it.
        raise FileError(
            "That looks like a binary file rather than a CSV. If it came out of "
            "Excel, save it as CSV UTF-8 or as .xlsx.")

    first = next((line for line in text.splitlines() if line.strip()), "")
    if not first:
        raise FileError("That file is empty — there is not even a header row in it.")

    sheet = Sheet(kind="csv", header=[])
    header: Optional[List[str]] = None
    for number, cells in enumerate(
            csv.reader(io.StringIO(text), delimiter=_delimiter(first)), start=1):
        if header is None:
            if not any(c.strip() for c in cells):
                continue                       # a blank line above the headings
            header = [fold(c) for c in cells]
            sheet.header = header
            continue
        if any(c.strip() for c in cells):
            sheet.rows.append((number, [c.strip() for c in cells]))
    if header is None:
        raise FileError("That file is empty — there is not even a header row in it.")
    return sheet


def _header_of(tab) -> List[str]:
    for row in tab.iter_rows(min_row=1, max_row=1, values_only=True):
        return [fold(c) for c in (row or ())]
    return []


def _pick(book, want: str):
    """Which tab of the workbook the decisions are on.

    The first one, unless it has no `id` column and another does. A workbook
    that has been through somebody's own working copy often has a scratch tab in
    front of the one that came out of the export, and picking the first sheet
    blindly would answer "no id column" about a file that plainly has one.
    """
    tabs = list(book.worksheets)
    if not tabs:
        return None
    for tab in tabs:
        if want in _header_of(tab):
            return tab
    return tabs[0]


def _from_xlsx(payload: bytes) -> Sheet:
    try:
        import openpyxl
    except ImportError:
        # Imported here rather than at the top, deliberately. openpyxl is the
        # only dependency this feature adds, and a machine that missed it should
        # still take the CSV the export actually produces — not fail to start,
        # and not fail on the import line of routes that have nothing to do with
        # Excel.
        raise FileError(
            "This service cannot read .xlsx files — openpyxl is not installed on "
            "it. Save the sheet as CSV and upload that, and tell whoever deploys "
            "this that requirements.txt has a line the machine has not picked up.")

    try:
        book = openpyxl.load_workbook(
            io.BytesIO(payload),
            # The values, not the formulas: somebody who typed =IF(...) into the
            # decision column meant the answer it worked out. read_only keeps a
            # long sheet from being built into objects a cell at a time.
            read_only=True, data_only=True,
        )
    except Exception as failure:               # openpyxl raises several kinds
        raise FileError(f"That .xlsx file could not be opened: {failure}")

    tab = _pick(book, "id")
    if tab is None:
        book.close()
        raise FileError("That workbook has no sheets in it.")

    sheet = Sheet(kind="xlsx", header=[], tab=tab.title)
    header: Optional[List[str]] = None
    for number, row in enumerate(tab.iter_rows(values_only=True), start=1):
        cells = [_text(c) for c in (row or ())]
        if header is None:
            if not any(cells):
                continue
            header = [fold(c) for c in cells]
            sheet.header = header
            continue
        if any(cells):
            sheet.rows.append((number, cells))
    book.close()
    if header is None:
        raise FileError("That sheet is empty — there is not even a header row in it.")
    return sheet


def read(payload: bytes) -> Sheet:
    """Whatever was uploaded, as a header and numbered rows — or FileError."""
    if not payload:
        raise FileError("That upload is empty — no file arrived at all.")
    if len(payload) > MAX_BYTES:
        # No size quoted, because the caller stops reading at the cap and does
        # not know the real one. A number that is always exactly the limit reads
        # as a measurement and is not one.
        raise FileError(
            f"That file is larger than the {MAX_BYTES // (1024 * 1024)}MB this "
            f"will read. A register of decisions is not that big, so it is "
            f"almost certainly the wrong file.")
    if payload.startswith(_OLE2):
        raise FileError(
            "That is an old-format .xls workbook, which this cannot read. Open "
            "it in Excel and save it as .xlsx or as CSV.")
    if payload.startswith(_ZIP):
        return _from_xlsx(payload)
    return _from_csv(payload)


def cell(cells: List[str], index: Optional[int]) -> str:
    """One column of one row, when the row may be shorter than the header.

    Excel drops trailing empty cells, so a row whose last few columns are blank
    comes back short and an index into it raises. Absent and empty are the same
    answer here, and both mean "no decision on this row".
    """
    if index is None or index < 0 or index >= len(cells):
        return ""
    return (cells[index] or "").strip()
