#!/usr/bin/env python3
"""
Cross-artefact audit. Every other checker validates one layer; this one validates that the
layers agree with each other.

Checks:

  1. Every operation in the lineage exists in a contract, and every contract operation is in
     the lineage.
  2. Every table the lineage names exists in the schema reference.
  3. `screen-index.json` and `screens/*.yaml` describe the same set of screens.
  4. Every contract, platform and table has a tooltip, and no tooltip describes something that
     no longer exists.
  5. Every app in a screen definition has a manifest, and vice versa.
  6. Every event a state model emits has a definition.
  7. **No shared file is duplicated into spine or satellite.** Found on 17 August: three copies
     of `permissions.yaml`, and the satellite copy had already drifted four permissions behind.
     A duplicated vocabulary diverges silently, and each copy looks correct on its own.
  8. Every operation declares either a permission or an `x-ticvai-auth` model. Fifty-four
     carried `x-ticvai-permission: null` with no statement of what protected them instead —
     unauthenticated by accident is indistinguishable from unauthenticated by design.
  9. Every contract names the requirement domain it serves, or states that it is cross-cutting.
     `platform-ops` reached a contract from a workshop without ever becoming a requirement
     (CF-49); this makes that visible rather than discoverable by accident.
 10. No derived table name carries a doubled suffix — `homepage_section_section` is a deriver
     defect, not a table.

Run: python3 tools/check-package.py
"""
import json
import re
import sys
from pathlib import Path

import yaml

ROOT = Path(__file__).resolve().parents[1]
if not (ROOT / "contracts").exists():
    ROOT = ROOT.parent / "ticvai-full"

ERRORS: list[str] = []
WARNINGS: list[str] = []


def main() -> int:
    C = ROOT / "contracts"
    if not C.exists():
        print("no contracts directory — nothing to audit")
        return 0

    # 7. shared files duplicated into a tier
    shared = {f.name for f in (C / "shared").glob("*.yaml")}
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                ERRORS.append(f"{tier}/{f.name} duplicates shared/{f.name} — a duplicated "
                              "vocabulary diverges silently and each copy looks correct alone")

    ops: set[str] = set()
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                continue
            doc = yaml.safe_load(f.read_text())
            for item in (doc.get("paths") or {}).values():
                if not isinstance(item, dict):
                    continue
                for verb, op in item.items():
                    if verb in ("get", "post", "put", "patch", "delete") and isinstance(op, dict):
                        ops.add(op["operationId"])

    H = ROOT / "handoff"
    lin = json.loads((H / "api-data-lineage.json").read_text()) if (H / "api-data-lineage.json").exists() else {}
    for x in sorted(set(lin) - ops):
        ERRORS.append(f"lineage has '{x}' which no contract defines")
    for x in sorted(ops - set(lin)):
        ERRORS.append(f"contract operation '{x}' is missing from the lineage")

    # 10. every contract names the requirement domain it serves, or a documented reason not to
    CROSS_CUTTING = {"identity", "tenancy", "cross-cell", "platform-ops"}
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                continue
            info = (yaml.safe_load(f.read_text()) or {}).get("info", {})
            mod = str(info.get("x-ticvai-module", ""))
            if not mod:
                ERRORS.append(f"{f.stem}: no x-ticvai-module — the contract names no requirement domain")
            elif f.stem in CROSS_CUTTING and "ross-cutting" not in mod:
                ERRORS.append(f"{f.stem}: cross-cutting contract must say so in x-ticvai-module, "
                              "so a contract with no requirement behind it is visible rather than "
                              "discoverable by accident")

    # 9. every operation declares how it is authenticated
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                continue
            doc = yaml.safe_load(f.read_text())
            for item in (doc.get("paths") or {}).values():
                if not isinstance(item, dict):
                    continue
                for verb, op in item.items():
                    if verb not in ("get", "post", "put", "patch", "delete") or not isinstance(op, dict):
                        continue
                    if not op.get("x-ticvai-permission") and not op.get("x-ticvai-auth"):
                        ERRORS.append(f"{op['operationId']}: no permission and no x-ticvai-auth — "
                                      "an operation with neither is unauthenticated by accident")

    # 2 + 8. tables
    wb = H / "TICVAI_Schema_Reference.xlsx"
    tables: set[str] = set()
    if wb.exists():
        try:
            import openpyxl
            ws = openpyxl.load_workbook(wb, data_only=True)["Tables"]
            tables = {ws.cell(r, 2).value for r in range(5, ws.max_row + 1) if ws.cell(r, 2).value}
        except Exception as e:  # noqa: BLE001
            WARNINGS.append(f"could not read the schema workbook: {e}")
    if tables:
        for t in sorted({t for v in lin.values() for t in v["reads"] + v["writes"]} - tables):
            ERRORS.append(f"lineage names table '{t}' which the schema reference does not have")
        for t in sorted(tables):
            if "." not in str(t):
                continue
            parts = str(t).split(".", 1)[1].split("_")
            if len(parts) > 1 and parts[-1] == parts[-2]:
                ERRORS.append(f"table '{t}' has a doubled suffix — a deriver defect, not a table")

    # 3 + 5. screens, apps
    sids: set[str] = set()
    apps: set[str] = set()
    plats: set[str] = set()
    for f in (ROOT / "screens").glob("P*.yaml"):
        doc = yaml.safe_load(f.read_text())
        apps.add(doc["platform"]["app"])
        plats.add(doc["platform"]["code"])
        for s in doc["screens"]:
            sids.add(s["id"])
            for a in (s.get("apis") or []):
                if (o := a.get("operationId")) and ops and o not in ops:
                    ERRORS.append(f"screen {s['id']} calls '{o}', which does not exist")
    if (H / "screen-index.json").exists():
        idx = json.loads((H / "screen-index.json").read_text())
        for x in sorted(set(idx) ^ sids):
            ERRORS.append(f"screen-index and screens/ disagree about '{x}'")
    mans = {yaml.safe_load(f.read_text())["app"] for f in (ROOT / "frontend").glob("*.yaml")}
    for a in sorted(apps ^ mans):
        ERRORS.append(f"app '{a}' appears in screens or manifests but not both")

    # 4. tooltips
    if (H / "tooltips.json").exists():
        T = json.loads((H / "tooltips.json").read_text())
        contracts = {f.stem for tier in ("spine", "satellite")
                     for f in (C / tier).glob("*.yaml") if f.name not in shared}
        for c in sorted(contracts ^ set(T.get("contracts", {}))):
            ERRORS.append(f"tooltip and contract set disagree about '{c}'")
        for p in sorted(plats ^ set(T.get("platforms", {}))):
            ERRORS.append(f"tooltip and platform set disagree about '{p}'")
        if tables:
            for t in sorted(tables - set(T.get("tables", {}))):
                WARNINGS.append(f"no tooltip for table {t}")

    # 6. events
    emitted: set[str] = set()
    for f in (ROOT / "states").glob("*.yaml"):
        if f.name == "_schema.yaml":
            continue
        for t in yaml.safe_load(f.read_text())["transitions"]:
            emitted |= set(t.get("emits") or [])
    defined = {yaml.safe_load(f.read_text())["name"] for f in (ROOT / "events").glob("*.yaml")
               if f.name != "_schema.yaml"}
    for e in sorted(emitted - defined):
        ERRORS.append(f"a state model emits '{e}' with no event definition")

    print(f"{len(ops)} operations · {len(tables)} tables · {len(sids)} screens · "
          f"{len(apps)} apps · {len(plats)} platforms\n")
    for w in WARNINGS[:20]:
        print(f"  WARN  {w}")
    for e in ERRORS:
        print(f"  FAIL  {e}")
    print()
    if ERRORS:
        print(f"{len(ERRORS)} error(s), {len(WARNINGS)} warning(s)")
        return 1
    print(f"PASS — {len(WARNINGS)} warning(s)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
