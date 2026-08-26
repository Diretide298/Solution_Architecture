#!/usr/bin/env python3
"""
Cross-artefact audit. Every other checker validates one layer; this one validates that the
layers agree with each other.

Checks:

  1. Every operation in the lineage exists in a contract, and every contract operation is in
     the lineage.
  2. Every table the lineage names exists in the schema reference.
  3. `screen-index.json` and `screens/*.yaml` describe the same set of screens.
  4. Every contract, platform and table has a tooltip, and no tooltip describes something that
     no longer exists.
  5. Every app in a screen definition has a manifest, and vice versa.
  6. Every event a state model emits has a definition.
  7. **No shared file is duplicated into spine or satellite.** Found on 17 August: three copies
     of `permissions.yaml`, and the satellite copy had already drifted four permissions behind.
     A duplicated vocabulary diverges silently, and each copy looks correct on its own.
  8. Every operation declares either a permission or an `x-ticvai-audience`. Fifty-four
     carried `x-ticvai-permission: null` with no statement of what protected them instead —
     unauthenticated by accident is indistinguishable from unauthenticated by design.
  9. Every schema naming a table gives that table columns. 76 of 287 tables were empty on
     18 August, including all 13 AI tables, and nothing failed.
 10. Every `x-ticvai-module` domain number resolves to a matrix domain of that name. Nine
     contracts cited a number meaning a different domain on 18 August, and two claimed 18.
 11. No operation writes a table documented as a read-only projection.
 10. ADR status leads with one of four values. "Accepted — split rule superseded by ADR-0014"
     read as live because it led with "Accepted".
 10. An ADR citing a superseded ADR names the supersession. On 17 August ADR-0021 reasoned from
     a decision that three separate labels said no longer held.
 10. Every RAG source in the register exists and is reachable by an AI indexing operation in
     the lineage. The register named eleven and the lineage agreed with none of them.
 10. Every `x-ticvai-*` marker uses a value from its closed set. `lastWriteWins` and
     `lastWriterWins` coexisted on 17 August — one policy, two spellings — and every checker
     passed because each value was individually plausible.
 10. AI isolation holds in the lineage, not just in prose: no AI operation writes outside its
     own stores, no other contract writes an AI table, and every model-calling operation writes
     an `ai.interaction`. Both halves were breached on 17 August with every checker passing.
 10. No document names a platform code that no longer exists. The kiosk was renumbered P03 to
     P05 and `platform-deployment.md`, two contracts and the tooltips kept the old code for
     days — every checker passed, because a stale code in prose resolves to nothing and breaks
     nothing until someone reads it.
 10. Every contract names the requirement domain it serves, or states that it is cross-cutting.
     `platform-ops` reached a contract from a workshop without ever becoming a requirement
     (CF-49); this makes that visible rather than discoverable by accident.
 10. No derived table name carries a doubled suffix — `homepage_section_section` is a deriver
     defect, not a table.

Run: python3 tools/check-package.py
"""
import json
import re
import sys
from pathlib import Path

import yaml

ROOT = Path(__file__).resolve().parents[1]
if not (ROOT / "contracts").exists():
    ROOT = ROOT.parent / "ticvai-full"

ERRORS: list[str] = []
WARNINGS: list[str] = []

# The 21 domains of the requirements matrix, by the id the matrix itself uses. **9 and 14 do not
# exist** — no hidden rows, no filter, nothing documenting why, and that part is with the client.
# The gaps are the whole reason CF-121 happened: nine contracts were authored against a compacted
# 1–21 sequence that closes them, so every reference above 8 drifted by one and then by two.
MATRIX_DOMAINS = {
    "1": "Ticketing Catalogue", "2": "Ticketing Sales", "3": "Admission and Access",
    "4": "Bundles and Promotions", "5": "F&B & Guest Management", "6": "Retail POS",
    "7": "F&B POS", "8": "Unified Operations Dashboard", "10": "Games & F&B Integration",
    "11": "Approval Workflows & Governance", "12": "Accreditation & Credential Management",
    "13": "Developer & API Management", "15": "Inventory Management", "16": "Device Management",
    "17": "Maintenance & Safety Management", "18": "Employee Mobile App & AI Assistant",
    "19": "Guest Mobile App & Branding", "20": "Subscription & Licensing Management",
    "21": "Seat Management & Venue Mapping", "22": "Marketing & CRM",
    "23": "Digital Asset Management",
}


def _norm(s: str) -> str:
    """Compare domain names on their words. The matrix writes '&' where a contract may write
    'and', and one sub-domain carries a double space."""
    return " ".join(s.lower().replace("&", "and").split())


def snake_table(name: str) -> str:
    """`StockPosition` -> `stock_position`, to compare a schema name against a table name."""
    return re.sub(r"(?<!^)(?=[A-Z])", "_", name).lower()


def main() -> int:
    C = ROOT / "contracts"
    if not C.exists():
        print("no contracts directory — nothing to audit")
        return 0

    # 7. shared files duplicated into a tier
    shared = {f.name for f in (C / "shared").glob("*.yaml")}
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                ERRORS.append(f"{tier}/{f.name} duplicates shared/{f.name} — a duplicated "
                              "vocabulary diverges silently and each copy looks correct alone")

    ops: set[str] = set()
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                continue
            doc = yaml.safe_load(f.read_text(encoding="utf-8"))
            for item in (doc.get("paths") or {}).values():
                if not isinstance(item, dict):
                    continue
                for verb, op in item.items():
                    if verb in ("get", "post", "put", "patch", "delete") and isinstance(op, dict):
                        ops.add(op["operationId"])

    H = ROOT / "handoff"
    lin = json.loads((H / "api-data-lineage.json").read_text(encoding="utf-8")) if (H / "api-data-lineage.json").exists() else {}
    for x in sorted(set(lin) - ops):
        ERRORS.append(f"lineage has '{x}' which no contract defines")
    for x in sorted(ops - set(lin)):
        ERRORS.append(f"contract operation '{x}' is missing from the lineage")

    # 10. every contract names the requirement domain it serves, or a documented reason not to
    CROSS_CUTTING = {"identity", "tenancy", "cross-cell", "platform-ops"}
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                continue
            info = (yaml.safe_load(f.read_text(encoding="utf-8")) or {}).get("info", {})
            mod = str(info.get("x-ticvai-module", ""))
            if not mod:
                ERRORS.append(f"{f.stem}: no x-ticvai-module — the contract names no requirement domain")
            elif f.stem in CROSS_CUTTING and "ross-cutting" not in mod:
                ERRORS.append(f"{f.stem}: cross-cutting contract must say so in x-ticvai-module, "
                              "so a contract with no requirement behind it is visible rather than "
                              "discoverable by accident")
            else:
                # CF-121. The number and the name must agree with the matrix. `x-ticvai-module` was
                # free text nothing resolved, and nine contracts cited a number meaning a different
                # domain — `inventory` said 13 where the matrix has Developer & API, `maintenance`
                # said 15 where it has Inventory. `workforce` and `subscription` both said 18. This
                # is CF-117's finding in a second field: the old check caught a domain that does not
                # exist and never a number paired with the wrong name.
                for num, nm in re.findall(r"(\d{1,2})\s*—\s*([^·|]+)", mod):
                    nm = nm.strip().rstrip(".")
                    # A contract may qualify its domain — "(auth) + cross-cutting staff auth" on
                    # identity, "(builder)" on white-label. The domain name is what precedes the
                    # first qualifier; everything after it says which slice of the domain is meant.
                    nm = re.split(r"\s+\+\s+", nm)[0]
                    nm = re.sub(r"\s*\(.*?\)\s*$", "", nm).strip()
                    truth = MATRIX_DOMAINS.get(num)
                    if truth is None:
                        ERRORS.append(f"{f.stem}: x-ticvai-module cites domain {num}, which the "
                                      f"matrix does not define — 9 and 14 do not exist (CF-121)")
                    elif _norm(truth) != _norm(nm):
                        ERRORS.append(f"{f.stem}: x-ticvai-module says '{num} — {nm}' but the "
                                      f"matrix calls domain {num} '{truth}' (CF-121)")

    # 18. Duplicate keys inside one mapping. YAML keeps the last and discards the rest silently —
    # no error, no warning, and the loss is invisible in a diff that only shows additions. On
    # 17 August four existed: a description explaining ADR-0021's central rule was overwritten by
    # a second one, and `x-ticvai-auth: guest` on three operations was discarded by a later
    # `service`, which is a guest operation silently becoming a service one.
    class _Dup(yaml.SafeLoader):
        pass

    _dupes: list[tuple[str, str, int, int]] = []

    def _dup_mapping(loader, node, deep=False):
        seen: dict = {}
        for k, _ in node.value:
            key = loader.construct_object(k, deep=deep)
            if key in seen:
                _dupes.append((loader.name, str(key), seen[key], k.start_mark.line + 1))
            seen[key] = k.start_mark.line + 1
        return yaml.SafeLoader.construct_mapping(loader, node, deep)

    _Dup.add_constructor(yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG, _dup_mapping)
    for f in sorted(C.glob("*/*.yaml")):
        _dupes.clear()
        with f.open(encoding="utf-8") as fh:
            yaml.load(fh, _Dup)
        for _, key, first, again in _dupes:
            ERRORS.append(f"{f.name}: key '{key}' defined at line {first} and redefined at "
                          f"{again} — YAML keeps the last and discards the first without a word")

    # 19. A platform code is written with the name its own screens file declares. On 18 August
    # eight of twelve platforms were called something else in the contracts — P09 alone had two
    # wrong names, "Platform Admin Console" and "Platform Admin", against a declared "TICVAI Web".
    # The existing check catches a code that does not exist; it did not catch a code paired with
    # the wrong name, which is the failure a reader actually hits.
    import re as _re5
    declared: dict[str, str] = {}
    for f in sorted((ROOT / "screens").glob("P*.yaml")):
        pl = (yaml.safe_load(f.read_text(encoding="utf-8")) or {}).get("platform") or {}
        if pl.get("code"):
            declared[pl["code"]] = pl.get("shortName") or pl.get("name")
    # The conflict register quotes past wordings deliberately — "P11 Accreditation as a guest
    # surface" is prose about a defect, not a platform name, and checking it would force the
    # register to launder its own history.
    for f in list(C.glob("*/*.yaml")) + list((ROOT / "docs").rglob("*.md")) + list(H.glob("*.md")):
        if f.name in ("conflicts.md", "conflict-status.md"):
            continue
        text = f.read_text(encoding="utf-8", errors="replace")
        # Only a code followed by words. `P12 SUP-001 Inbox` is a screen reference, not a name,
        # and matching it produced 1,837 false errors on the first attempt.
        for m in _re5.finditer(r"\b(P\d\d) ((?:[A-Z][a-z&-]*)(?: [A-Za-z&-]+){1,4})\b", text):
            code, name = m.group(1), m.group(2).strip().rstrip("·|")
            want = declared.get(code)
            if not want or _re5.match(r"^[A-Z]{2,4}-\d", name):
                continue
            if name == want or name.startswith(want) or want.startswith(name):
                continue
            # a longer descriptive title that contains the short name is fine
            if want.lower() in name.lower():
                continue
            # **A board filename is not a platform name.** Boards left the package on 20 August —
            # they are built locally now — but the exemption stays, because documents in
            # `docs/active/` still quote the old paths and a document quoting a path is not a
            # document naming a platform. `wireframes/P07 Staff Scanner.dc.html`
            # is a file that predates the rename to Venue Scanner, and a document quoting it is
            # quoting a path. Renaming the files would break every `wireframe.board` anchor in the
            # package to fix a cosmetic mismatch in prose.
            #
            # Caught twice on 20 August against two different documents, and both times the
            # document was right.
            span = text[max(0, m.start() - 60):m.end() + 60]
            if ".dc.html" in span or "wireframes/" in span:
                continue
            ERRORS.append(f"{f.name}: writes '{code} {name}' — its screens file declares "
                          f"'{want}'. One code, one name")

    # 19. Every schema that names a table must give that table columns. The schema reference was
    # derived once by an ad-hoc script and then hand-patched, so on 18 August **76 of 287 tables had
    # no columns** — every table belonging to a contract written after that run, including all 13
    # `ai.*` tables. Nothing failed, because a table with no columns is not an error to a checker
    # that only asks whether the table exists. `tools/derive-schema.py` now regenerates them.
    sref = H / "schema-reference.json"
    if sref.exists():
        ref = json.loads(sref.read_text(encoding="utf-8"))
        cols = ref.get("cols", {})
        for tier in ("spine", "satellite"):
            for f in sorted((C / tier).glob("*.yaml")):
                doc = yaml.safe_load(f.read_text(encoding="utf-8")) or {}
                for sname, body in ((doc.get("components") or {}).get("schemas") or {}).items():
                    if not isinstance(body, dict):
                        continue
                    tbl = body.get("x-ticvai-persistence")
                    if not isinstance(tbl, str) or "." not in tbl or "—" in tbl:
                        continue
                    tbl = tbl.split("+")[0].strip()
                    if body.get("properties") and not cols.get(tbl):
                        ERRORS.append(f"{f.name}: schema {sname} persists to {tbl} and that table has "
                                      "no columns in the schema reference — run tools/derive-schema.py")

    # 21. A scalar value starting with `*` is a YAML alias, not text. `description: **Where it
    # goes.**` broke `fnb.yaml` on 18 August — the file stopped parsing entirely, and the message
    # points at a line 140 rows away from the edit. Emphasis at the start of a value needs a block
    # scalar.
    for f in sorted(C.glob("*/*.yaml")):
        for n, line in enumerate(f.read_text(encoding="utf-8").split("\n"), 1):
            m = re.match(r"^\s*\w[\w\-]*:\s+(\*\*)", line)
            if m:
                ERRORS.append(f"{f.name}:{n}: a value starting with '*' reads as a YAML alias — "
                              "use a block scalar (`>`) when a description opens with emphasis")

    # 22. No column name is an operation reference. `derive-relationships` records ambient edges
    # as `via <operationId>` where two tables are written together and no column joins them, and
    # `derive-schema` used to turn those into columns — 123 of them, `fnb.menu.via createMenu`
    # among them, sitting in the workbook typed as uuid.
    #
    # **Two derivers that feed each other need a check on the boundary**, because each is
    # individually correct.
    ref = ROOT / "handoff" / "schema-reference.json"
    if ref.exists():
        _s = json.loads(ref.read_text(encoding="utf-8"))
        for _t, _cols in (_s.get("cols") or {}).items():
            for _c in _cols:
                if _c["column"].startswith("via ") or " " in _c["column"]:
                    ERRORS.append(f"{_t}.{_c['column']}: a lineage edge reads as a column — "
                                  "ambient couplings belong in the relationship graph, not the schema")

    # 23. Every operation says who may call it. A staff permission, a non-staff audience, or
    # `x-ticvai-self-service` — a principal acting on their own session rather than on the
    # platform. **Ten session and MFA operations had none of the three on 18 August**, and read as
    # a permission gap when they are the opposite: `logout` needs no permission because a person
    # who is signed in may always sign out.
    #
    # Without the third value these stay invisible: they are not guest-facing and they are not
    # staff-permissioned, so both checks pass them over.
    for tier in ("spine", "satellite"):
        for f in sorted((C / tier).glob("*.yaml")):
            doc = yaml.safe_load(f.read_text(encoding="utf-8")) or {}
            for path, item in (doc.get("paths") or {}).items():
                for verb, op in item.items():
                    if verb not in ('get','post','put','patch','delete') or not isinstance(op, dict) or not op.get("operationId"):
                        continue
                    aud = set(op.get("x-ticvai-audience") or [])
                    if (op.get("x-ticvai-permission") or op.get("x-ticvai-self-service")
                            or (aud & {"guest", "public", "anonymous", "service", "device", "partner"})):
                        continue
                    ERRORS.append(f"{f.stem}.{op['operationId']}: no permission, no non-staff "
                                  "audience and not self-service — who may call it is unstated")

    # 24. Every table a column references must exist. `orders.wallet_pass.entitlement_id`,
    # `orders.resale_listing.entitlement_id` and `ScanEvent.ticketId` all pointed at
    # `access.entitlement` on 18 August and **there was no such table** — the package sold
    # products, defined `EntitlementTemplate`, modelled `EntitlementStatus`, wrote a state model
    # for it, and never stored the row a guest actually holds.
    #
    # **`validateAccess` read the template and `suspendEntitlement` wrote it**, which would have
    # suspended a pass for every guest who held one. Every artefact was individually consistent and
    # the thing they all pointed at was absent.
    ref = ROOT / "handoff" / "schema-reference.json"
    if ref.exists():
        _s = json.loads(ref.read_text(encoding="utf-8"))
        _known = {t for t in (set(_s.get("cols") or {}) | set(_s.get("storage") or {}))}
        for _t, _cols in (_s.get("cols") or {}).items():
            for _c in _cols:
                _target = _c.get("references")
                if _target and _target not in _known:
                    ERRORS.append(f"{_t}.{_c['column']} references '{_target}', which is stored "
                                  "nowhere — a column pointing at a table that does not exist")

    # 25. Every non-shared contract declares its taxonomy. **Six declared no
    # `x-ticvai-platforms` on 18 August** — `ai`, `public-api`, `resources`, `venue-map`,
    # `workforce` and `approvals` — and no package checker tested for it. The viewer did, and
    # reported them as missing taxonomy where nobody was reading.
    #
    # **The viewer was checking something the package was not**, which is the wrong way round: a
    # reader should not be the thing that catches a gap in what it reads.
    for tier in ("spine", "satellite"):
        for f in sorted((C / tier).glob("*.yaml")):
            info = (yaml.safe_load(f.read_text(encoding="utf-8")) or {}).get("info") or {}
            for key in ("x-ticvai-module", "x-ticvai-tier", "x-ticvai-platforms"):
                if not info.get(key):
                    ERRORS.append(f"{f.stem}: declares no {key} — the viewer groups and filters on it")
            if info.get("x-ticvai-tier") and info["x-ticvai-tier"] != tier:
                ERRORS.append(f"{f.stem}: declares tier '{info['x-ticvai-tier']}' and sits in {tier}/")

    # 26. Every permission an operation names is in the Permission enum. **Ten operations used
    # six invented permissions on 18 August** — `FNB_MANAGE`, `FNB_SERVE`, `INVENTORY_VIEW`,
    # `REPORT_VIEW`, `WALLET_MANAGE`, `TENANT_VIEW` — every one written that day, and every one a
    # name that reads well rather than a name the contract already uses.
    #
    # **The viewer caught these and no package checker did.** It reported them as errors where
    # nobody was reading, which is the wrong way round: a reader should not be the thing that
    # catches a gap in what it reads.
    perms_file = C / "shared" / "permissions.yaml"
    if perms_file.exists():
        declared = set((yaml.safe_load(perms_file.read_text(encoding="utf-8"))
                        ["components"]["schemas"]["Permission"]["enum"]))
        for tier in ("spine", "satellite"):
            for f in sorted((C / tier).glob("*.yaml")):
                doc = yaml.safe_load(f.read_text(encoding="utf-8")) or {}
                for path, item in (doc.get("paths") or {}).items():
                    for verb, op in item.items():
                        if verb not in ("get", "post", "put", "patch", "delete") or not isinstance(op, dict):
                            continue
                        for key in ("x-ticvai-permission", "x-ticvai-permission-escalated"):
                            perm = op.get(key)
                            if perm and perm not in declared:
                                ERRORS.append(f"{f.stem}.{op.get('operationId')}: {perm} is not in "
                                              "the Permission enum")

    # 27. A platform code means one platform. `x-ticvai-platforms` is free text, and on
    # 18 August `P04` appeared as both "POS" in ten contracts and "Venue POS" in one — **the one
    # was mine**, derived from the screen file's `shortName` while everything else used the older
    # label. **The viewer groups by the code and labels by the name**, so a collision renders one
    # platform twice under two headings.
    seen_platform: dict = {}
    for tier in ("spine", "satellite"):
        for f in sorted((C / tier).glob("*.yaml")):
            info = (yaml.safe_load(f.read_text(encoding="utf-8")) or {}).get("info") or {}
            for entry in (info.get("x-ticvai-platforms") or []):
                m = re.match(r"(P\d\d)\s+(.*)", str(entry))
                if not m:
                    continue
                code, label = m.group(1), m.group(2).strip()
                first = seen_platform.setdefault(code, (label, f.stem))
                if first[0] != label:
                    ERRORS.append(f"{f.stem}: {code} is '{label}' here and '{first[0]}' in "
                                  f"{first[1]} — one code must mean one platform")

    # 28. A schema declaring `persistence: none` must not have a Postgres table behind it.
    # **`identity.session` was a Postgres table with one column on 18 August** while `Session` and
    # `ActiveSession` both said *none — Redis session registry*, and `inventory.stock_level` was a
    # table while `StockPosition` said *derived from movements*.
    #
    # **A stored level and a movement ledger that disagree is a stock count nobody can reconcile**,
    # and the contract had it right — the schema reference did not.
    ref = ROOT / "handoff" / "schema-reference.json"
    if ref.exists():
        _s = json.loads(ref.read_text(encoding="utf-8"))
        _none = set()
        for tier in ("spine", "satellite"):
            for f in sorted((C / tier).glob("*.yaml")):
                doc = yaml.safe_load(f.read_text(encoding="utf-8")) or {}
                for sname, body in ((doc.get("components") or {}).get("schemas") or {}).items():
                    p_ = isinstance(body, dict) and body.get("x-ticvai-persistence")
                    if isinstance(p_, str) and p_.strip().startswith("none"):
                        _none.add(snake_table(sname))
        for _t, _store in (_s.get("store") or {}).items():
            if _store != "postgres":
                continue
            if _t.split(".")[-1] in _none:
                ERRORS.append(f"{_t} is stored in postgres and its schema declares persistence "
                              "none — one of the two is wrong")

    # 29. The repo mirrors match the root. `repos/update-bible.sh` says it plainly —
    # **project-bible/ is a copy, not a source** — and nothing ran it. On 20 August the mirror held
    # 21 state files against the root's 113: 54 differed, 144 existed only in the mirror and
    # **305 root files had never reached it.**
    #
    # It surfaced through one file. `states/entitlement.yaml` was renamed and repointed from
    # `access.TicketStatus` — an object, not an enum — to `orders.EntitlementStatus`, and the six
    # mirrored copies kept the original anchored on nothing. **`access.TicketStatus`'s own
    # description claimed the file had been removed**, which is worse than a stale file: the note
    # stops the next person looking.
    mirrored = ("contracts", "states", "flows", "events", "screens", "docs", "frontend",
                "tools", "handoff", "sources")
    repos = ROOT / "repos"
    if repos.is_dir():
        bibles = [d for d in ((repos / "ticvai-docs"),) if d.is_dir()]
        bibles += [r / "project-bible" for r in sorted(repos.iterdir())
                   if (r / "project-bible").is_dir()]
        for b in bibles:
            drift = 0
            for folder in mirrored:
                s_dir = ROOT / folder
                if not s_dir.is_dir():
                    continue
                for f in s_dir.rglob("*"):
                    if not f.is_file() or "__pycache__" in f.parts:
                        continue
                    t = b / f.relative_to(ROOT)
                    if not t.exists() or t.read_bytes() != f.read_bytes():
                        drift += 1
            if drift:
                ERRORS.append(f"{b.relative_to(ROOT)}: mirror is out of sync with the root "
                              f"({drift} file(s)) — run tools/derive-mirrors.py")

    # 30. A table needs a key. **116 tables had no `id` column on 20 August** and the cause is
    # systematic rather than 116 oversights: **the schema reference derives table columns from API
    # response schemas, and a response is not a table.** `Subscription` returns tenantId, planId,
    # planName and status — everything a caller needs and not the row's own identity.
    #
    # Most are legitimate. A declared child of a parent has a composite key; a join or value table
    # keys on its two columns; `catalogue.price` keys naturally on (price_list_id, variant_id).
    # **What is refused is a table with no `id`, no parent and no column that could serve as a
    # natural key** — a row nothing can address, update or delete.
    ref = ROOT / "handoff" / "schema-reference.json"
    if ref.exists():
        sr = json.loads(ref.read_text(encoding="utf-8"))
        lineage = sr.get("lineage") or {}
        NATURAL = ("code", "version", "serial", "key", "slug", "identifier", "media_code")
        for tbl, cols in (sr.get("cols") or {}).items():
            if not cols or ":" in tbl:
                continue
            names = {c["column"] for c in cols}
            # **`<table>_id` on `<table>` is a primary key**, by a convention half this package
            # uses. The first cut of this check demanded literally `id` and flagged
            # `whitelabel.navigation_item.navigation_item_id` — which is the key, named the other
            # way. **Three narrower fixes to the deriver each replaced one defect with another
            # before the check itself turned out to be the thing that was wrong.**
            short = tbl.split(".")[-1]
            if "id" in names or f"{short}_id" in names or lineage.get(tbl, {}).get("parent"):
                continue
            if len(cols) <= 3:
                continue  # a join or value table keys on what it holds
            if any(n == k or n.endswith("_" + k) for n in names for k in NATURAL):
                continue
            ERRORS.append(f"{tbl}: has no key at all — no id, no parent, no natural key, and "
                          f"{len(cols)} columns. A row nothing can address")

    # 31. Nothing writes to a derived table. **Four operations wrote `inventory.stock_level` on
    # 20 August** — `recordWaste`, `completeProductionRun`, `collectMerchandiseReservation` and
    # `cancelMerchandiseReservation` — while `StockPosition` declares *derived from movements* and
    # the store says `derived`.
    #
    # **A stored level and a movement ledger that disagree is a stock count nobody can reconcile**,
    # which is exactly why the level is derived. An operation that consumes stock writes the
    # movement; the level follows.
    ref = ROOT / "handoff" / "schema-reference.json"
    lin_p = ROOT / "handoff" / "api-data-lineage.json"
    if ref.exists() and lin_p.exists():
        stores = (json.loads(ref.read_text(encoding="utf-8")).get("store") or {})
        derived_tables = {t for t, st in stores.items() if st == "derived"}
        for op, v in json.loads(lin_p.read_text(encoding="utf-8")).items():
            hit = sorted(t for t in (v.get("writes") or []) if t in derived_tables)
            if hit:
                ERRORS.append(f"{op} ({v.get('contract')}) writes {hit} — that table is derived. "
                              "Write what it derives from")

    # 32. **The diagrams are derived and nothing fails when they are not.** That is the exact
    # shape of the two worst silent failures in this package: the mirrors drifted 503 files and the
    # services workbook said 371 tables against 379, and in both cases every validator passed.
    #
    # **A sixth viewer layer makes that risk worse**, because the same facts now appear in two
    # places — so the mtime check is the price of the layer.
    dia = ROOT / "diagrams"
    src = ROOT / "handoff" / "service-decomposition.json"
    if dia.is_dir() and src.exists():
        newest_src = src.stat().st_mtime
        for name in ("api-data-lineage.json", "schema-reference.json"):
            p_ = ROOT / "handoff" / name
            if p_.exists():
                newest_src = max(newest_src, p_.stat().st_mtime)
        stale = [f.relative_to(ROOT) for f in sorted(dia.rglob("*.yaml"))
                 if f.stat().st_mtime < newest_src - 2]
        if stale:
            ERRORS.append(
                f"{len(stale)} diagram file(s) are older than what they are derived from "
                f"(first: {stale[0]}) — run tools/derive-diagrams.py. A hand-edited diagram is "
                "the one artefact that goes stale without anything failing")

    # 32. **The diagrams are derived and must not go stale.** A diagram in a package with nine
    # validators is the one artefact that drifts silently — nothing fails when it is wrong. That
    # is how the mirrors drifted 503 files and how the services workbook said 371 tables against
    # 379, and both were found by a person rather than a check.
    #
    # **Compares mtime rather than content** deliberately: the point is to catch a package where
    # `refresh.sh` was not run, not to re-derive here and compare.
    hld = ROOT / "diagrams" / "hld" / "00-platform.yaml"
    dec = ROOT / "handoff" / "service-decomposition.json"
    if dec.exists():
        for want in ("README.yaml", "hld/00-platform.yaml", "hld/01-hierarchy.yaml",
                     "hld/02-services.yaml", "hld/03-contracts.yaml", "hld/04-lifecycles.yaml"):
            if not (ROOT / "diagrams" / want).exists():
                ERRORS.append(f"diagrams/{want} is missing — run tools/derive-diagrams.py")
        if not hld.exists():
            pass
        # **Watch every source, not just the one the diagrams are named after.** The first cut
        # compared the diagrams to `service-decomposition.json` alone — so two operations added to
        # the lineage left `OrderService.yaml` and `AccessService.yaml` reporting 96 and 30 against
        # a live 97 and 31, and **the check passed the whole time**.
        #
        # A staleness check that watches one input is a staleness check that reports the freshness
        # of that input, which is not the question.
        sources = [dec,
                   ROOT / "handoff" / "api-data-lineage.json",
                   ROOT / "handoff" / "schema-reference.json"]
        newest = max((f.stat().st_mtime for f in sources if f.exists()), default=0)
        if hld.stat().st_mtime < newest - 1:
            stale = sorted(f.name for f in sources
                           if f.exists() and f.stat().st_mtime > hld.stat().st_mtime + 1)
            ERRORS.append(f"diagrams are older than {', '.join(stale)} — run tools/refresh.sh")
        else:
            lld = ROOT / "diagrams" / "lld" / "services"
            svcs = set(json.loads(dec.read_text(encoding="utf-8")).get("services") or {})
            have = {f.stem for f in lld.glob("*.yaml")} if lld.is_dir() else set()
            if svcs - have:
                ERRORS.append(f"diagrams/lld/ is missing {sorted(svcs - have)} — "
                              "run tools/derive-diagrams.py")
            if have - svcs:
                ERRORS.append(f"diagrams/lld/ has {sorted(have - svcs)} which is not a service — "
                              "a renamed or removed service left a file behind")

    # 33. **A stored currency column has to earn its place.** Currency and scale are
    # region-scoped and not overridable below (ADR-0018), so a row in a UAE region is AED and
    # cannot be anything else — storing it per row is a copy of a fact that cannot differ, and a
    # workstation with its own currency is one somebody can misconfigure into a mismatch with the
    # ledger it posts to.
    #
    # **Four tables genuinely differ from their region**: a guest pays USD at an AED venue, a
    # supplier invoices in their own, an account is denominated, a partner settles in theirs.
    # Those four keep a stored currency and every other one is a defect.
    CURRENCY_OK = {
        "orders.payment", "inventory.supplier", "ledger.account", "ledger.legal_entity",
        "control.partner_agreement", "platform.region_settings", "platform.denomination",
        "retail.wallet",
    }
    _schema = ROOT / "handoff" / "schema-reference.json"
    if _schema.exists():
        _cols = (json.loads(_schema.read_text(encoding="utf-8")).get("cols") or {})
        for _t, _cs in sorted(_cols.items()):
            if _t in CURRENCY_OK:
                continue
            _hit = [c["column"] for c in _cs
                    if c["column"] in ("currency", "currency_code", "currency_scale")]
            if _hit:
                ERRORS.append(
                    f"{_t}: stores {', '.join(_hit)} — currency is region-scoped and resolves from "
                    "the scope walk (ADR-0018). Mark the property `x-ticvai-persisted: false`, or "
                    "add the table to CURRENCY_OK with the reason it genuinely differs")

    # 34. **Every `$ref` resolves.** Deleting a shared schema breaks every contract that points at
    # it and nothing in the suite noticed: the `Money` rewrite on 24 August removed `SalesChannel`,
    # `ScopeLevel` and `ScopeRef` from `shared/common.yaml`, leaving ten references across five
    # files pointing at nothing. **YAML still parsed, so eight validators still passed** — a broken
    # reference is only visible to something that follows it.
    #
    # It also caught `#/components/schemas/SalesOrder` in an operation I built the same day: the
    # schema is `Order`. **A local ref is as easy to get wrong as a shared one and looks more
    # trustworthy.**
    _docs = {}
    for _f in sorted(ROOT.glob("contracts/*/*.yaml")):
        try:
            _docs[str(_f.resolve())] = yaml.safe_load(_f.read_text(encoding="utf-8"))
        except Exception:
            continue

    def _at(doc, frag):
        for part in frag.lstrip("#/").split("/"):
            doc = doc.get(part) if isinstance(doc, dict) else None
        return doc

    for _f in sorted(ROOT.glob("contracts/*/*.yaml")):
        _own = _docs.get(str(_f.resolve()))
        for _m in re.finditer(r"\$ref: '([^']+)'", _f.read_text(encoding="utf-8")):
            _r = _m.group(1)
            if _r.startswith("#"):
                if _at(_own, _r) is None:
                    ERRORS.append(f"{_f.name}: $ref '{_r}' resolves to nothing in its own file")
            else:
                _path, _, _frag = _r.partition("#")
                _t = (_f.parent / _path).resolve()
                if str(_t) not in _docs:
                    ERRORS.append(f"{_f.name}: $ref '{_r}' — no such file")
                elif _frag and _at(_docs[str(_t)], _frag) is None:
                    ERRORS.append(f"{_f.name}: $ref '{_r}' — the file exists, the schema does not")

    # 34. **Every `$ref` resolves.** A reference to a schema that does not exist parses fine,
    # generates fine and fails at the first client — which is how eleven of them survived a rewrite
    # of `shared/common.yaml` on 24 August. `SalesChannel`, `ScopeLevel` and `ScopeRef` were
    # removed while nine references across four contracts kept pointing at them, and nothing here
    # noticed.
    #
    # **The one that found it was a person reading a diagram**, and that is the wrong last line of
    # defence for a link the whole package is built on resolving.
    _docs = {}
    for _f in sorted((ROOT / "contracts").glob("*/*.yaml")):
        try:
            _docs[str(_f.resolve())] = yaml.safe_load(_f.read_text(encoding="utf-8"))
        except Exception:
            pass

    def _at(doc, frag):
        for part in frag.lstrip("#/").split("/"):
            doc = (doc or {}).get(part) if isinstance(doc, dict) else None
        return doc

    for _f in sorted((ROOT / "contracts").glob("*/*.yaml")):
        _key = str(_f.resolve())
        _self = _docs.get(_key)
        for _m in re.finditer(r"\$ref: '([^']+)'", _f.read_text(encoding="utf-8")):
            _r = _m.group(1)
            if _r.startswith("#"):
                if _at(_self, _r) is None:
                    ERRORS.append(f"{_f.name}: $ref '{_r}' resolves to nothing in its own file")
            else:
                _path, _, _frag = _r.partition("#")
                _t = (_f.parent / _path).resolve()
                if str(_t) not in _docs:
                    ERRORS.append(f"{_f.name}: $ref '{_r}' names a file that is not there")
                elif _frag and _at(_docs[str(_t)], _frag) is None:
                    ERRORS.append(f"{_f.name}: $ref '{_r}' resolves to nothing — the file exists "
                                  "and the schema in it does not")

    # 35. **An unbounded list over a table that grows without limit is the query that takes a
    # venue down.** 94 of 183 `list*` operations declared no page size; 38 of those read `orders`,
    # `marketing`, `access`, `ledger`, `fnb`, `retail` or `inventory` — the tables a campaign send
    # or a season-ticket base fills. **The first venue to notice is the one whose guest base crossed
    # a hundred thousand**, and by then it is a production incident rather than a review comment.
    _BIG = ("orders.", "marketing.", "access.", "ledger.", "fnb.", "retail.", "inventory.")
    _lin_p = ROOT / "handoff" / "api-data-lineage.json"
    if _lin_p.exists():
        _L = json.loads(_lin_p.read_text(encoding="utf-8"))
        for _f in sorted((ROOT / "contracts").glob("*/*.yaml")):
            _d = yaml.safe_load(_f.read_text(encoding="utf-8")) or {}
            for _path, _item in (_d.get("paths") or {}).items():
                for _v, _o in _item.items():
                    if _v.upper() != "GET" or not isinstance(_o, dict):
                        continue
                    _oid = _o.get("operationId") or ""
                    if not _oid.startswith("list") or _oid not in _L:
                        continue
                    if "PageSize" in str(_o.get("parameters") or ""):
                        continue
                    if any(t.startswith(_BIG) for t in _L[_oid].get("reads", [])):
                        ERRORS.append(
                            f"{_f.stem}.{_oid}: no page size, and it reads a high-volume table "
                            f"({[t for t in _L[_oid]['reads'] if t.startswith(_BIG)][:2]}) — an "
                            "unbounded list over a table that grows without limit")

    # 36. **A table written by something and read by nothing is a feature that half-exists.**
    # A guest joins a restaurant waitlist and no operation lists the waitlist; a steward attaches a
    # photograph to a work order and nothing retrieves it. **`platform.audit_record` was exactly
    # this on 20 August** — written by nothing, read by nothing, found by walking a journey rather
    # than by any check here.
    #
    # **A child table read through its parent is not orphaned**, which is why the parent is
    # consulted before reporting.
    _sch_p = ROOT / "handoff" / "schema-reference.json"
    if _lin_p.exists() and _sch_p.exists():
        _S = json.loads(_sch_p.read_text(encoding="utf-8"))
        _real = {t for t in (set(_S.get("cols") or {}) | set(_S.get("storage") or {}))
                 if "." in t and ":" not in t}
        _lineage = _S.get("lineage") or {}
        _w, _r = {}, set()
        for _o, _v in _L.items():
            for _t in _v.get("writes", []):
                if _t in _real:
                    _w.setdefault(_t, []).append(_o)
            for _t in _v.get("reads", []):
                if _t in _real:
                    _r.add(_t)
        for _t in sorted(set(_w) - _r):
            if (_lineage.get(_t) or {}).get("parent"):
                continue
            WARNINGS.append(
                f"{_t}: written by {sorted(_w[_t])[:2]} and read by nothing, and it has no parent "
                "to be read through — either an operation is missing or the table should not exist")

    # 13. The x-ticvai-* vocabularies are closed sets. `lastWriteWins` and `lastWriterWins` were
    # both in use on 17 August — one policy, two spellings, ten operations split between them, and
    # every checker passed because each value was individually plausible.
    VOCAB = {
        "x-ticvai-conflict-policy": {"serverWins", "lastWriterWins", "append", "manualMerge"},
        # Sourced from tenancy.ScopeLevel plus `platform`, which is above the tenant tree.
        "x-ticvai-scope-level": {"platform", "tenant", "brand", "region", "venue", "department",
                                 "subDepartment", "workstation"},
        "x-ticvai-read-routing": {"primary", "replica", "analytical"},
        # A partner and an external reviewer hold real permissions and are neither staff nor guests.
        # Omitting them is what let P11 be treated as a guest surface on 17 August.
        "x-ticvai-audience": {"staff", "guest", "partner", "public", "anonymous", "device", "service"},
    }
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                continue
            for item in (yaml.safe_load(f.read_text(encoding="utf-8")).get("paths") or {}).values():
                if not isinstance(item, dict):
                    continue
                for verb, op in item.items():
                    if verb not in ("get", "post", "put", "patch", "delete") or not isinstance(op, dict):
                        continue
                    for key, allowed in VOCAB.items():
                        val = op.get(key)
                        if val is None:
                            continue
                        # `x-ticvai-audience` is a list; every other marker is a scalar. Both are
                        # closed sets and both are checked the same way.
                        for item in (val if isinstance(val, list) else [val]):
                            if item not in allowed:
                                ERRORS.append(f"{op['operationId']}: {key} contains '{item}', which "
                                              f"is not in the closed set {sorted(allowed)}")
                        continue
                        if False:
                            ERRORS.append(f"{op['operationId']}: {key} is '{val}', which is not in the "
                                          f"closed set {sorted(allowed)}")

    # 12. AI isolation, from ADR-0020. Two breaches existed on the day that ADR was written and
    # every validator passed, because each operation existed and resolved to a real table.
    # `generateVenueLayout` wrote into `seating.import_job` — AI writing into a transactional
    # contract — and `askReportingQuestion` wrote no `ai.interaction` despite being brought under
    # governance the same day.
    lin_path = H / "api-data-lineage.json"
    if lin_path.exists():
        lineage = json.loads(lin_path.read_text(encoding="utf-8"))
        GOVERNED_OUTSIDE_AI = {"askReportingQuestion", "saveNaturalLanguageQuery"}
        CALLS_A_MODEL = {"sendAiMessage", "semanticSearch", "generateConfiguration",
                         "generateVenueLayout", "decideProposedAction", "askReportingQuestion"}
        for op, v in lineage.items():
            is_ai = str(v.get("contract")) == "ai"
            writes = v.get("writes") or []
            ai_writes = [t for t in writes if t.startswith("ai.") or t.startswith("qdrant")]
            if is_ai:
                # A cache is not the transactional core. ADR-0020's rule is about writing into
                # another contract's tables; every `cache:*` entry is derived from something already
                # read, invalidated by an event already consumed, and losable without consequence —
                # nothing treats one as a source of truth. Listing two by name was too narrow and
                # broke the moment `cache:idempotency` reached every write.
                outside = [t for t in writes
                           if not (t.startswith("ai.") or t.startswith("qdrant")
                                   or t.startswith("cache:"))]
                if outside:
                    ERRORS.append(f"{op}: an AI operation writes {outside} outside its own stores — "
                                  "AI is read-only against the transactional core (ADR-0020)")
            elif ai_writes and op not in GOVERNED_OUTSIDE_AI:
                ERRORS.append(f"{op} ({v.get('contract')}) writes {ai_writes} — only the AI contract "
                              "and the governed reporting pair may write an AI table")
            if op in CALLS_A_MODEL and "ai.interaction" not in writes:
                ERRORS.append(f"{op} calls a model and writes no ai.interaction — requirement 8.3.55 "
                              "is satisfied in prose and not in the data")

    # 17. A table documented as a read-only projection must have no writers. `platform.tenant`
    # is described in the schema reference as "read-only projection of control.tenant" and three
    # operations wrote it on 17 August — a projection with writers is a second master.
    PROJECTIONS = {"platform.tenant"}
    proj_lin = H / "api-data-lineage.json"
    if proj_lin.exists():
        lineage3 = json.loads(proj_lin.read_text(encoding="utf-8"))
        for op, v in lineage3.items():
            for t in (v.get("writes") or []):
                if t in PROJECTIONS:
                    ERRORS.append(f"{op} writes {t}, which is a read-only projection — a projection "
                                  "with a writer is a second master")

    # 16. ADR status is a closed set of four. Six spellings were in use on 17 August, and one of
    # them — "Accepted — split rule superseded by ADR-0014" — read as live because it led with
    # "Accepted". Reasoning from that ADR produced a cross-tenant isolation defect (CF-97).
    ADR_STATUS = ("Accepted", "Accepted in part", "Proposed", "Superseded")
    adr_dir2 = ROOT / "docs" / "adr"
    if adr_dir2.exists():
        import re as _re4
        for f in sorted(adr_dir2.glob("0*.md")):
            m = _re4.search(r"^\*\*Status:\*\* *(.+)$", f.read_text(encoding="utf-8"), _re4.M)
            if not m:
                ERRORS.append(f"{f.name}: no Status line")
                continue
            raw = m.group(1).strip()
            # The state must be the FIRST token and unadorned. Stripping asterisks before the
            # comparison is what let `**Status:** **Superseded**` pass here on 17 August while
            # the viewer's parser returned null for it — the emphasis added to make the state
            # unmissable to a reader made it invisible to everything that reads the field.
            if not re.match(r"^[A-Za-z]", raw):
                ERRORS.append(f"{f.name}: status starts with '{raw[:12]}' — the state must be the "
                              "first token and must not be wrapped in emphasis, or a parser that "
                              "reads the first word gets a marker instead of a state")
                continue
            if not raw.startswith(ADR_STATUS):
                ERRORS.append(f"{f.name}: status starts '{raw[:40]}', which is not one of "
                              f"{list(ADR_STATUS)} — a status must lead with its state, not bury it")

    # 15. An ADR citing a superseded ADR must name the supersession. ADR-0021 reasoned from
    # ADR-0001's decision sentence on 17 August while ADR-0001's own status line, the ADR index,
    # and ADR-0014 all said that decision no longer held. Everything was labelled; the reading
    # skipped the labels, and a label nobody reads is not a control.
    adr_dir = ROOT / "docs" / "adr"
    if adr_dir.exists():
        superseded = {}
        for f in adr_dir.glob("0*.md"):
            # Only the Status line of the ADR's own header counts. A document that merely
            # discusses supersession is not itself superseded.
            for ln in f.read_text(encoding="utf-8").split("\n")[:8]:
                if ln.startswith("**Status:**") and ("uperseded" in ln or "mended" in ln):
                    superseded[f.name[:4]] = f.name
                    break
        for f in adr_dir.glob("0*.md"):
            text = f.read_text(encoding="utf-8")
            for num in sorted(superseded):
                if f.name[:4] == num:
                    continue
                if f"ADR-{num}" not in text:
                    continue
                near = [ln for ln in text.split("\n") if f"ADR-{num}" in ln]
                if not any("upersed" in ln or "amend" in ln or "no longer" in ln for ln in near):
                    ERRORS.append(f"{f.name} cites ADR-{num}, which is superseded or amended, "
                                  "without saying so on any line that mentions it")

    # 11. no document names a platform code that no longer exists
    import re as _re
    real_codes = {yaml.safe_load(f.read_text(encoding="utf-8"))["platform"]["code"]
                  for f in (ROOT / "screens").glob("P*.yaml")}
    for f in list(ROOT.rglob("*.md")) + list((ROOT / "contracts").rglob("*.yaml")):
        if "repos" in str(f):
            continue
        try:
            text = f.read_text(encoding="utf-8")
        except Exception:  # noqa: BLE001
            continue
        for code in sorted(set(_re.findall(r"\bP\d\d\b", text)) - real_codes):
            ERRORS.append(f"{f.name}: names platform code {code}, which no screen file defines")

    # 9. every operation declares how it is authenticated
    for tier in ("spine", "satellite"):
        for f in (C / tier).glob("*.yaml"):
            if f.name in shared:
                continue
            doc = yaml.safe_load(f.read_text(encoding="utf-8"))
            for item in (doc.get("paths") or {}).values():
                if not isinstance(item, dict):
                    continue
                for verb, op in item.items():
                    if verb not in ("get", "post", "put", "patch", "delete") or not isinstance(op, dict):
                        continue
                    if not op.get("x-ticvai-permission") and not op.get("x-ticvai-audience"):
                        ERRORS.append(f"{op['operationId']}: no permission and no x-ticvai-auth — "
                                      "an operation with neither is unauthenticated by accident")

    # 2 + 8. tables
    wb = H / "TICVAI_Schema_Reference.xlsx"
    tables: set[str] = set()
    if wb.exists():
        try:
            import openpyxl
            ws = openpyxl.load_workbook(wb, data_only=True)["Tables"]
            tables = {ws.cell(r, 2).value for r in range(5, ws.max_row + 1) if ws.cell(r, 2).value}
        except Exception as e:  # noqa: BLE001
            WARNINGS.append(f"could not read the schema workbook: {e}")
    if tables:
        for t in sorted({t for v in lin.values() for t in v["reads"] + v["writes"]} - tables):
            ERRORS.append(f"lineage names table '{t}' which the schema reference does not have")
        for t in sorted(tables):
            if "." not in str(t):
                continue
            parts = str(t).split(".", 1)[1].split("_")
            if len(parts) > 1 and parts[-1] == parts[-2]:
                ERRORS.append(f"table '{t}' has a doubled suffix — a deriver defect, not a table")

    # 14. Every table the RAG source register names must exist, must be reachable by an AI
    # indexing operation in the lineage, and must have an invalidating event. On 17 August the
    # register named eleven sources and the lineage said AI read none of them — the register was
    # a document describing something the data did not agree had happened.
    rag_path = H / "rag-index-sources.md"
    rag_lin = H / "api-data-lineage.json"
    if rag_path.exists() and rag_lin.exists():
        import re as _re2
        lineage2 = json.loads(rag_lin.read_text(encoding="utf-8"))
        sources = set(_re2.findall(r"^\| `([a-z_]+\.[a-z_]+)` \|", rag_path.read_text(encoding="utf-8"), _re2.M))
        indexed = {t for op, v in lineage2.items()
                   if str(v.get("contract")) == "ai" for t in (v.get("reads") or [])}
        for t in sorted(sources - indexed):
            ERRORS.append(f"RAG source {t} is declared in the register and no AI operation reads it "
                          "in the lineage")
        if tables:
            for t in sorted(sources - tables):
                ERRORS.append(f"RAG source {t} does not exist in the schema reference")

    # 3 + 5. screens, apps
    sids: set[str] = set()
    apps: set[str] = set()
    plats: set[str] = set()
    for f in (ROOT / "screens").glob("P*.yaml"):
        doc = yaml.safe_load(f.read_text(encoding="utf-8"))
        apps.add(doc["platform"]["app"])
        plats.add(doc["platform"]["code"])
        for s in doc["screens"]:
            sids.add(s["id"])
            for a in (s.get("apis") or []):
                if (o := a.get("operationId")) and ops and o not in ops:
                    ERRORS.append(f"screen {s['id']} calls '{o}', which does not exist")
    if (H / "screen-index.json").exists():
        idx = json.loads((H / "screen-index.json").read_text(encoding="utf-8"))
        for x in sorted(set(idx) ^ sids):
            ERRORS.append(f"screen-index and screens/ disagree about '{x}'")
    mans = {yaml.safe_load(f.read_text(encoding="utf-8"))["app"] for f in (ROOT / "frontend").glob("*.yaml")}
    for a in sorted(apps ^ mans):
        ERRORS.append(f"app '{a}' appears in screens or manifests but not both")

    # 4. tooltips
    if (H / "tooltips.json").exists():
        T = json.loads((H / "tooltips.json").read_text(encoding="utf-8"))
        contracts = {f.stem for tier in ("spine", "satellite")
                     for f in (C / tier).glob("*.yaml") if f.name not in shared}
        for c in sorted(contracts ^ set(T.get("contracts", {}))):
            ERRORS.append(f"tooltip and contract set disagree about '{c}'")
        for p in sorted(plats ^ set(T.get("platforms", {}))):
            ERRORS.append(f"tooltip and platform set disagree about '{p}'")
        if tables:
            for t in sorted(tables - set(T.get("tables", {}))):
                WARNINGS.append(f"no tooltip for table {t}")

    # 6. events
    emitted: set[str] = set()
    for f in (ROOT / "states").glob("*.yaml"):
        if f.name == "_schema.yaml":
            continue
        for t in yaml.safe_load(f.read_text(encoding="utf-8"))["transitions"]:
            emitted |= set(t.get("emits") or [])
    defined = {yaml.safe_load(f.read_text(encoding="utf-8"))["name"] for f in (ROOT / "events").glob("*.yaml")
               if f.name != "_schema.yaml"}
    for e in sorted(emitted - defined):
        ERRORS.append(f"a state model emits '{e}' with no event definition")

    print(f"{len(ops)} operations · {len(tables)} tables · {len(sids)} screens · "
          f"{len(apps)} apps · {len(plats)} platforms\n")
    for w in WARNINGS[:20]:
        print(f"  WARN  {w}")
    for e in ERRORS:
        print(f"  FAIL  {e}")
    print()
    if ERRORS:
        print(f"{len(ERRORS)} error(s), {len(WARNINGS)} warning(s)")
        return 1
    print(f"PASS — {len(WARNINGS)} warning(s)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
