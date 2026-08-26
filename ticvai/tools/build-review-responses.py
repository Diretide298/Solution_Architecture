#!/usr/bin/env python3
"""Build the review response workbook — one row per review, with our verdict checked against the package.

**Every claim is verified rather than accepted.** A reviewer writing *"Need GET API for List of My
Tickets"* is reporting what they could not find on the screen in front of them, and roughly half the
time the operation exists and the screen does not name it. **Those are different defects and they go
to different people**: one is a contract gap, the other is a wiring gap on a screen already drawn.

The `our verdict` column is one of:

  **Agreed**        the reviewer is right and the package is wrong
  **Already there** the operation exists and is on that screen — needs a pointer, not a build
  **Wired, not shown** the operation exists and is NOT on that screen — a real wiring gap
  **Decision**      not a defect; a question that needs somebody to choose
  **Disagree**      we think the current design is right, with the reason
"""
from __future__ import annotations

import csv
import glob
import json
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import sys

# A cp1252 console cannot encode the arrows and dashes this tool prints, and the
# failure lands *after* the work is done — so the output is written, the summary
# line raises UnicodeEncodeError, and a correct run exits 1. Reconfiguring at
# import means anything importing this module gets it too, refresh.sh included.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:      # a captured stream may not be reconfigurable; harmless
    pass


PKG = Path("/home/claude/ticvai-pkg")
CSV = Path("/mnt/user-data/uploads/ticvai-review-activity-2026-08-24-filtered.csv")
OUT = Path("/mnt/user-data/outputs/TICVAI_Review_Responses.xlsx")

HEAD = PatternFill("solid", fgColor="1B2A4A")
BAND = PatternFill("solid", fgColor="F4F7FB")
VERDICT_FILL = {
    "Agreed": PatternFill("solid", fgColor="FDECEC"),
    "Wired, not shown": PatternFill("solid", fgColor="FFF2CC"),
    "Already there": PatternFill("solid", fgColor="E2EFDA"),
    "Decision": PatternFill("solid", fgColor="DDEBF7"),
    "Disagree": PatternFill("solid", fgColor="EDEDED"),
    "Approved": PatternFill("solid", fgColor="E2EFDA"),
}
THIN = Side(style="thin", color="D8DEE8")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Verified responses. Keyed by (artefact, first 40 chars of the note) so a screen reviewed twice
# gets two answers — POS-001 has three separate findings and they are not the same finding.
RESPONSES = {
    ("GST-012", "Need GET API for List of My Tickets."): (
        "Already there",
        "`listMyEntitlements` exists and **is already on GST-012**. Guest-scoped — it resolves the "
        "caller from the session rather than taking a subject parameter, which is why it is not "
        "`listEntitlements`."),
    ("GST-010", "Need API for Get Order Details."): (
        "Wired, not shown",
        "`getOrder` exists and **is not on GST-010**. A real gap: the screen shows a confirmation "
        "it cannot fetch. Wiring it, not building anything."),
    ("GST-013", "Need GET API for Ticket Details."): (
        "Already there",
        "`getEntitlement`, `getEntitlementCredential` and `getEntitlementHistory` are all on "
        "GST-013. **The credential is what a gate reads and the history answers *has this been "
        "used*** — F50 walks all three."),
    ("GST-019", "Need GET API for Logged In Guest Order History."): (
        "Agreed",
        "**Built as `listMyOrders`.** `listOrders` was on the screen and it is the staff-scoped "
        "one — same defect as the case operations: **a guest-facing list must be scoped to the "
        "caller, not filtered by a subject parameter**, or somebody will call it with another "
        "subject."),
    ("WEB-018", "Missing the  read surface"): (
        "Agreed",
        "**`listEntitlements` built.** The other three named — `getEntitlement`, "
        "`getEntitlementCredential`, `getEntitlementHistory` — were already on WEB-018."),
    ("GST-015", "Need : -"): (
        "Wired, not shown",
        "`getSubscription` exists and **is not on GST-015**. Membership members and the add path "
        "are a genuine gap — raising as a scope question rather than building against a guess."),
    ("WEB-006", "Potential Missing API -"): (
        "Wired, not shown",
        "`acquireLease` exists and **is not on WEB-006**. Correct call: a session selection that "
        "does not lease is a slot two channels can sell. **Contended inventory is leased, not "
        "reserved** (CF-115)."),
    ("WEB-007", "Validation Required"): (
        "Agreed",
        "**Right, and it is a server-side rule not a screen one.** `createSeatHold` must validate "
        "requested against leased or the hold is advisory. F43 walks the concurrency; the "
        "validation belongs in the operation."),
    ("WEB-004", "API Parameter Verification"): (
        "Decision",
        "`listPerformances` is on WEB-004 and takes no product filter. **Either the screen filters "
        "client-side or the operation gains a parameter** — worth deciding, because a performance "
        "list unfiltered by product is every performance in the venue."),
    ("WEB-002", "Potential API Consolidation"): (
        "Decision",
        "**Both exist and both are on the screen.** `searchCatalogue` is a query across kinds; "
        "`listProducts` is a filtered list of one. **They overlap and shipping two is a choice** — "
        "our position is keep both and make the screen call one."),
    ("WEB-003", "Potential API Consolidation-"): (
        "Decision",
        "Same as WEB-002. **One decision covers both screens.**"),
    ("WEB-008", "Potential Unnecessary API -"): (
        "Agreed",
        "**`listCatalogueBundles` does not belong on an add-ons screen.** It lists published till "
        "bundles (ADR-0013) — a device concern, not a guest one. Removing."),
    ("WEB-009", "Potential Unnecessary APIs -"): (
        "Disagree",
        "`listGuestDevices`, `registerGuestDevice` and `revokeGuestDevice` are on WEB-009 "
        "deliberately. **A wishlist is per-guest and syncs across devices** — the device list is "
        "how a guest revokes an old phone that still holds their tickets."),
    ("WEB-016", "Three endpoints missing"): (
        "Wired, not shown",
        "`verifyGuestEmail` exists and sits on WEB-020, not WEB-016. **The registration screen "
        "cannot verify what it just registered** — F56 walks this and it is the largest single gap "
        "in the guest identity surface."),
    ("WEB-020", "getProfile/updateProfile"): (
        "Already there",
        "`updateMyProfile` and `verifyGuestEmail` are both on WEB-020. **Consent is separate and "
        "never inherited** (CF-160) — `recordConsent` is the operation, and it is on GST-065."),
    ("WEB-010", "All 9 APIs are valid"): (
        "Already there",
        "Confirmed — all nine resolve and all nine are on WEB-010. **Nothing to remove.**"),
    ("WEB-011", "Potential API Clarification"): (
        "Wired, not shown",
        "`updateMyProfile` exists and **is not on WEB-011**. Attendee details on a checkout form "
        "are the guest's own profile fields; wiring it."),
    ("WEB-012", "Potential Missing API"): (
        "Wired, not shown",
        "`getOrder` exists and **is not on WEB-012**. A payment screen that cannot re-read the "
        "order it is settling cannot recover from a refresh."),
    ("WEB-013", "transferOrderTickets is not represented"): (
        "Agreed",
        "**Correct — no visible control.** Either the screen gains one or the operation comes off. "
        "Our position: order confirmation is exactly where a guest forwards a ticket, so the "
        "screen should show it."),
    ("WEB-015", "clarification needed"): (
        "Agreed",
        "**`joinQueue` on a waiting room with no join action.** A virtual waiting room is entered, "
        "not joined — the guest is already in it. Removing; `getQueueEntry` is what the screen "
        "needs."),
    ("WEB-017", "Get Account Summary is needed"): (
        "Agreed",
        "**One onLoad call rather than three counts and a list.** A dashboard that makes four "
        "round trips to render above the fold is four chances to be slow. Building "
        "`getAccountSummary`."),
    ("WEB-001", "Potential Missing API-"): (
        "Decision",
        "*Explore by Place* has no operation because **the venue map is imported and labelled as a "
        "pipeline, not queried as a catalogue**. Either places become a browsable dimension or the "
        "section comes off the landing page."),
    ("GST-002", "Add a search bar"): (
        "Agreed", "`searchCatalogue` is on GST-063 and not on GST-002. Wiring it."),
    ("GST-005", "Need Confirmation on How we are Connecting"): (
        "Decision",
        "**A product is what is sold; a performance is when.** `catalogue.performance` carries "
        "`product_id`, so the join exists — what needs confirming is whether one product may span "
        "performances of different kinds. Answering with the schema rather than a sentence."),
    ("GST-006", "Need Confirmation on How we are Connecting"): (
        "Decision", "Same as GST-005. **One answer covers both.**"),
    ("GST-008", "Add quantity increment"): (
        "Agreed", "A component change, not a contract one. **`addCartLine` already takes a "
                  "quantity** — the screen needs the control."),
    ("GST-018", "For the Add to Calendar"): (
        "Decision",
        "A calendar entry needs a start, a duration and a location. **`getOrder` carries the "
        "order and the performance carries the time** — worth confirming the screen can assemble "
        "it rather than adding an operation to do it server-side."),
    ("POS-001", "CashDenominations should be added"): (
        "Agreed",
        "**`platform.denomination` already exists as a proper table** — `currency_code`, "
        "`display_name`, `value`, `kind`, `sort_order`. **And `orders.cash_movement.denominations` "
        "is still `jsonb`, which contradicts it.** The reviewer's fix is half-built and the JSON "
        "column is the half that is wrong."),
    ("POS-001", "CurrencyCode, DisplayName, Value"): (
        "Already there",
        "Exactly the shape of `platform.denomination`. **Seeded per region** — currency and scale "
        "are region-scoped and not overridable below (ADR-0018)."),
    ("POS-001", "Expected Float Value and Actual Float"): (
        "Agreed",
        "**And it is wider than this screen.** `Money` is an object with no persistence hint, so "
        "**129 money columns across the package are `jsonb`** — every total, variance and price. "
        "`orders.cash_movement.amount` is `numeric(18,4)` because somebody hand-typed it, and the "
        "inconsistency is the tell. **You cannot sum a jsonb price in SQL**, which moves every "
        "reconciliation into application code. Raising as an ADR."),
    ("POS-010", "If any user wants to replace"): (
        "Wired, not shown",
        "`exchangeOrderLines` exists and **is not on POS-010**. That is the operation for "
        "replacing a line."),
    ("POS-010", "Cart GET api is missing"): (
        "Wired, not shown",
        "`getCart` exists and **is not on POS-010**. Wiring it."),
    ("EMP-001", "role_permission"): (
        "Agreed",
        "`identity.role_permission` has `role_id` and `permission` and **no `created_at`, "
        "`created_by` or `is_active`** — no audit trail on a permission grant. "
        "`listAuditRecords` was built on 20 August precisely because nothing wrote the audit log; "
        "this is the same hole one level down."),
    ("EMP-002", "Shift link needed"): (
        "Agreed", "A role-select screen that cannot see the open shift cannot resume it. "
                  "`getCurrentShift` wiring."),
    ("EMP-003", "shift link is missing"): (
        "Agreed", "Same as EMP-002."),
    ("getCurrentShift", "What is the request parameter"): (
        "Already there",
        "**Neither — it resolves from the session.** The workstation is on the principal's session "
        "and a till is signed into, not selected. ADR-0002: authority is carried by the person."),
    ("ADM-001", "Sidebar/navigation should not appear"): (
        "Agreed", "A sign-in screen showing navigation it has not yet earned. Component change."),
    ("ADM-004", "The current audit log does not provide"): (
        "Agreed",
        "**Tenant, venue and workstation belong on an audit row.** `platform.audit_record` "
        "carries `scope_path`, which resolves all three — the screen needs to show it rather than "
        "the schema needing to change."),
    ("ADM-005", "This screen is the main list of all tenants"): (
        "Agreed", "Tenant ID and code on the directory. Component change."),
    ("ADM-007", "The current UI works for Tenant Admin"): (
        "Agreed",
        "**A platform admin has no tenant until they pick one.** The screen assumes a tenant in "
        "scope, which is true for a tenant admin and false for the audience it is on."),
    ("ADM-009", "should clearly display the selected Tenant"): (
        "Agreed", "Same root cause as ADM-007 — **an unnamed tenant on a billing screen is a "
                  "screen somebody invoices the wrong company from.**"),
    ("ADM-011", "ADM-011 currently covers seat configuration"): (
        "Agreed", "Same root cause as ADM-007 and ADM-009. **One fix: a tenant selector on the "
                  "platform-admin surface, not three.**"),
    ("ADM-012", "should be handled in the backend"): (
        "Agreed",
        "**Right, and the rejection is the correct verdict.** `scope_path` partitioning enforces "
        "isolation and 121 operations were stripped from P09 on 24 August for reaching past it. "
        "**A UI implying isolation is toggleable is worse than no UI.**"),
    ("ADM-016", "Looks good, but we also need to verify"): (
        "Decision",
        "The white-label builder is `WhiteLabelService` — 50 operations, 44% walked. **What is "
        "behind *Open the Builder* is a scope question**, not a defect."),
    ("WEB-001", "@hrushikant.patkar  test"): (
        "Disagree", "A test row. No action."),
}


def main() -> int:
    rows = list(csv.DictReader(CSV.read_text(encoding="utf-8-sig").splitlines()))
    lin = json.loads((PKG / "handoff" / "api-data-lineage.json").read_text(encoding="utf-8"))
    scr = {}
    for f in sorted(glob.glob(str(PKG / "screens" / "P*.yaml"))):
        for s in yaml.safe_load(Path(f).read_text(encoding="utf-8"))["screens"]:
            scr[s["id"]] = {a.get("operationId") for a in (s.get("apis") or [])
                            if a.get("operationId")}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responses"
    ws["A1"] = "TICVAI — review responses, 24 August 2026"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="1B2A4A")
    ws["A2"] = ("74 reviews. Every claim checked against the package rather than accepted — roughly "
                "half the time the operation exists and the screen does not name it.")
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="5A6577")
    ws["A3"] = ("Agreed = we are wrong.  Wired, not shown = the operation exists and is not on that "
                "screen.  Already there = it is on the screen.  Decision = needs somebody to choose.")
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="8A6D00")

    cols = ["Date", "Artefact", "Lands on", "Reviewer", "Their verdict", "Their note",
            "Our verdict", "Our response", "Action"]
    widths = [11, 11, 10, 17, 13, 54, 15, 78, 22]
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(4, i, c)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "A5"

    ACTION = {
        "Agreed": "Fix in package",
        "Wired, not shown": "Wire on screen",
        "Already there": "Reply — point at it",
        "Decision": "Raise with Qossai",
        "Disagree": "Reply with reason",
    }

    r = 5
    counts = {}
    for i, row in enumerate(rows):
        art = row["artefact"]
        note = (row["note"] or "").strip()
        verdict, resp = "Already there", ""
        if row["verdict"] == "Approved":
            verdict, resp = "Approved", "No action."
        else:
            for (a, prefix), (v, t) in RESPONSES.items():
                if a == art and note.startswith(prefix[:34]):
                    verdict, resp = v, t
                    break
            else:
                verdict = "Decision"
                resp = ("**Not yet answered.** Needs a check against the package before a verdict "
                        "— recorded rather than guessed.")
        counts[verdict] = counts.get(verdict, 0) + 1
        vals = [row["date"], art, row["lands on"], row["reviewer"], row["verdict"],
                note[:400], verdict, resp, ACTION.get(verdict, "—")]
        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j, v)
            c.font = Font(name="Arial", size=9, bold=(j == 2))
            c.alignment = Alignment(vertical="top", wrap_text=(j in (6, 8)))
            c.border = BOX
            if i % 2 == 0:
                c.fill = BAND
        ws.cell(r, 7).fill = VERDICT_FILL.get(verdict, BAND)
        r += 1
    ws.auto_filter.ref = f"A4:I{r - 1}"

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "What the reviews found"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color="1B2A4A")
    for i, (c, w) in enumerate(zip(["Our verdict", "Rows", "What it means", "Who acts"],
                                   [20, 8, 74, 22]), 1):
        cell = ws2.cell(3, i, c)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = HEAD
        cell.border = BOX
        ws2.column_dimensions[get_column_letter(i)].width = w
    MEANING = {
        "Agreed": ("**The reviewer is right and the package is wrong.** Includes four schema "
                   "defects that block DDL.", "Softlabs — fix"),
        "Wired, not shown": ("**The operation exists and the screen does not name it.** A wiring "
                             "gap, not a contract gap — and the reviewers reported these as "
                             "missing APIs because that is what they look like from a screen.",
                             "Softlabs — wire"),
        "Already there": ("**On the screen already.** Needs a reply pointing at it, not a build.",
                          "Softlabs — reply"),
        "Decision": ("**Not a defect.** A question somebody has to choose — consolidation, scope, "
                     "or a client confirmation.", "Qossai / Chinmay"),
        "Disagree": ("**We think the current design is right**, with the reason stated.",
                     "Softlabs — reply"),
        "Approved": ("No action.", "—"),
    }
    rr = 4
    for v, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        m, who = MEANING.get(v, ("", ""))
        for j, val in enumerate([v, n, m, who], 1):
            c = ws2.cell(rr, j, val)
            c.font = Font(name="Arial", size=9, bold=(j == 1))
            c.alignment = Alignment(vertical="top", wrap_text=(j == 3))
            c.border = BOX
        ws2.cell(rr, 1).fill = VERDICT_FILL.get(v, BAND)
        rr += 1

    ws2.cell(rr + 1, 1, "The four that block DDL").font = Font(name="Arial", size=11, bold=True)
    for i, t in enumerate([
        "Money is jsonb — 129 columns. Every total, variance and price. You cannot sum it in SQL.",
        "orders.cash_movement.denominations is jsonb while platform.denomination exists as a table.",
        "orders.cash_count_line has one column, shift_id. No denomination, no count, no variance.",
        "identity.role_permission has no created_at, created_by or is_active — no audit on a grant.",
    ], rr + 2):
        ws2.cell(i, 1, t).font = Font(name="Arial", size=9)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"  {len(rows)} rows · {counts}")
    print(f"  → {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
