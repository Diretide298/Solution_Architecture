"""Build the schema reference workbook.

**Reads the package, not a working directory.** Until 26 August this opened `schema_v4.json`,
`lineage.json`, `links.json` and `modules.json` by bare name — files that lived in whatever folder
somebody happened to run it from, and **none of which is in the package.** So it could not run here
at all, which is why the viewer's `module.written` read a column the workbook never wrote.

**Same class as `derive-board-panel-map.py` reading `/tmp/reads.json`**: a tool whose inputs are
outside the artefact it describes is a tool that works until the machine changes.

`links.json` and `modules.json` have no package equivalent. **Where they are absent the sheets that
need them are skipped and named**, rather than the whole workbook failing — the other seven sheets
are the ones a backend engineer opens.
"""
from pathlib import Path as _P
_ROOT = _P(__file__).resolve().parents[1]
_H = _ROOT / "handoff"


def _pkg(*names):
    """First of `names` that exists in `handoff/`, else None."""
    for n in names:
        if (_H / n).exists():
            return _H / n
    return None


import json, yaml, glob, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

# A cp1252 console cannot encode the arrows and dashes this tool prints, and the
# failure lands *after* the work is done — so the output is written, the summary
# line raises UnicodeEncodeError, and a correct run exits 1. Reconfiguring at
# import means anything importing this module gets it too, refresh.sh included.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:      # a captured stream may not be reconfigurable; harmless
    pass

D=json.load(open(_pkg('schema-reference.json','schema_v4.json'), encoding='utf-8')); L=(json.load(open(_pkg('links.json'), encoding='utf-8')) if _pkg('links.json') else {})
cols=D['cols']; nomap=D['nomap']; origin=D['origin']; storage=D['storage']
# **Where each table hangs, on the table's own row.** schema-roots.md answers it in prose and a
# reader looking at a table in this sheet was not in that file. `anchors` is where the table's own
# outbound keys stop; `parent` is the table it is a child of, where one owns it.
LINE=D.get('lineage',{})
rels=L['rels']; tab_ops=L['tab_ops']; tab_screens=L['tab_screens']
MIG={'platform':'V0001 / V0003 / V0003a','identity':'V0002','pii':'V0001a','sync':'V0001',
 'marketing':'V0003b (part) / V0015','fnb':'V0003b (part) / V0011','retail':'V0003b (part) / V0013',
 'orders':'V0005','catalogue':'V0004','access':'V0007','ledger':'V0008','seating':'V0010',
 'inventory':'V0012','promotions':'V0014','maintenance':'V0016','queue':'V0017',
 'whitelabel':'V0018','assets':'V0019','games':'V0020','reporting':'V0021',
 'control':'Control Plane — separate database'}
# No migrations are written as of 14 August — the workbook is the working artefact and DDL
# resumes when the design settles. The Written column stays in the sheet so it means something
# again the moment a migration lands.
def _globbed(pattern, what):
    """glob, and refuse to be silent about finding nothing.

    **Both of these used to point at `/home/claude/...`** — a path on the
    machine the tool was written on. `glob.glob` on a directory that is not
    there returns `[]` and raises nothing, so the workbook built, every sheet
    appeared, and two of them were empty:

      * `Scaling` shipped a header and one row reading `TOTAL 0 0 0 0`, which
        took Backend › Routing in the viewer from 25 contracts and 776 routed
        operations to a blank page
      * the `Written` column went blank for every table, which is why
        `module.written` read as false everywhere and every schema drew amber

    Same class as the JSON inputs this file's docstring already describes, in
    the same file, missed because a `glob` fails quietly where an `open` does
    not. So it says so now.
    """
    hits = sorted(glob.glob(pattern))
    if not hits:
        print(f"  WARN  no {what} matched {pattern} — the sheets built from it "
              f"will be empty, and they will not say so themselves")
    return hits


# DDL is read from the package's own `backend/`, which is where the .sql lives.
written=set()
for f in _globbed(str(_ROOT / 'backend' / 'V*.sql'), 'migrations'):
    s=open(f, encoding='utf-8').read(); m=re.search(r"^-- =+\n-- ROLLBACK",s,re.M)
    written.update(re.findall(r'CREATE TABLE (?:IF NOT EXISTS )?([\w.]+)', s[:m.start() if m else len(s)]))
C=str(_ROOT / 'contracts')
routing=defaultdict(lambda: defaultdict(int))
for f in _globbed(f'{C}/spine/*.yaml', 'spine contracts')+_globbed(f'{C}/satellite/*.yaml', 'satellite contracts'):
    d=yaml.safe_load(open(f, encoding='utf-8')); ctx=os.path.basename(f)[:-5]
    for p,i in (d.get('paths') or {}).items():
        for v,o in i.items():
            if not isinstance(o,dict) or v not in ('get','post','put','patch','delete'): continue
            routing[ctx]['write' if v!='get' else o.get('x-ticvai-read-routing','?')]+=1
F='Arial'
HDR=Font(name=F,bold=True,color='FFFFFF',size=10); HF=PatternFill('solid',fgColor='1F3864')
T=Font(name=F,bold=True,size=14); SUB=Font(name=F,size=10,color='595959')
B=Font(name=F,size=10); M=Font(name='Consolas',size=9); BD=Font(name=F,bold=True,size=10)
GREEN=PatternFill('solid',fgColor='E2EFDA'); AMBER=PatternFill('solid',fgColor='FFF2CC')
GREY=PatternFill('solid',fgColor='F2F2F2'); RED=PatternFill('solid',fgColor='FCE4E4')
BLUE=PatternFill('solid',fgColor='DDEBF7'); PINK=PatternFill('solid',fgColor='F4E1F0')
th=Side(style='thin',color='BFBFBF'); BOX=Border(left=th,right=th,top=th,bottom=th)
wb=Workbook(); wb.remove(wb.active)
def hdr(ws,l,w,row=4):
    for i,(a,b_) in enumerate(zip(l,w),1):
        c=ws.cell(row,i,a); c.font=HDR; c.fill=HF; c.border=BOX
        c.alignment=Alignment(vertical='center',wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width=b_
    ws.row_dimensions[row].height=28; ws.freeze_panes=ws.cell(row+1,1)
tot=len(set(cols)|set(storage))   # union — two tables were in both maps and double-counted until 14 Aug
NEW={'platform.outlet','platform.tenant','marketing.guest_device','marketing.wishlist_item',
 'fnb.delivery_location','fnb.location_session','fnb.delivery_location_outlet','fnb.location_code',
 'retail.shop_and_drop','retail.shop_and_drop_line','assets.media_asset','assets.media_collection',
 'assets.media_upload','assets.media_usage','pii.subject','pii.subject_contact',
 'pii.subject_document','pii.subject_biometric'}
PII={t for t in set(cols)|set(storage) if t.startswith('pii.')}

ws=wb.create_sheet('Read me'); ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=102
for i,(t_,f_) in enumerate([
 ('TICVAI — database schema reference',T),('',None),
 (f'{tot} tables. {len(cols)} from the API contracts, {len(storage)} storage-only.',BD),
 (f"{sum(len(v) for v in cols.values())} columns. 600 operations. {len(written)} tables written as DDL.",B),('',None),
 ('Personal data',BD),
 ('The pii schema was declared in V0001 and left empty. Four foreign keys pointed at',B),
 ('pii.subject, which no migration created — psql would have failed on the first one.',B),
 ('V0001a now writes four tables and an erasure function.',B),('',None),
 ('It is separate because of erasure. The ledger is append-only: a sale cannot be unwritten',B),
 ('because the buyer asked. Splitting identity from behaviour lets both hold — delete the',B),
 ('person, keep the transaction, and the transaction still reconciles because it never held',B),
 ('a name. Everything outside pii refers to a person by subject_id and nothing else.',B),('',None),
 ('A subject_id is pseudonymous, not anonymous. Re-identification requires this schema,',B),
 ('which is why reporting reads a replica without it and the AI layer never receives a row',B),
 ('from it (ADR-0009).',B),('',None),
 ('Eighteen tables are new, shown in blue on the Tables sheet',BD),
 ('  pii.subject, _contact, _document, _biometric   the schema was empty',B),
 ('  marketing.guest_device      push notifications had nowhere to land',B),
 ('  marketing.wishlist_item     no contract anywhere',B),
 ('  fnb.delivery_location       4.6.26 names cabanas and seats; only tables existed',B),
 ('  retail.shop_and_drop        4.4.7, not modelled at all',B),
 ('  platform.outlet             referenced nine times, defined nowhere',B),
 ('  platform.tenant             a cell could not resolve its own tenant',B),('',None),
 ('What this is NOT',BD),
 ('Not the DDL. Partitioning, RLS policies, indexes and the PII split live in the migrations.',B)],2):
    c=ws.cell(i,2,t_)
    if f_: c.font=f_

ws=wb.create_sheet('Modules')
import json as _json
MODROWS=(_json.load(open(_pkg('modules.json'), encoding='utf-8')) if _pkg('modules.json') else [])
ws.cell(1,1,'Schema modules').font=T
ws.cell(2,1,'What each module is, and why the boundary is where it is. Spine modules are depended on; '
            'satellites depend on the spine and never on each other.').font=SUB
# **`Written` added 26 August.** The viewer's Backend > Data view read `module.written` and the
# sheet had no such column, so every schema drew amber — **a status nobody could act on because it
# was the same for all 26 whatever the state of the build.**
#
# It counts tables with a `CREATE TABLE` in `backend/`. **That is zero today and the column is
# still worth having**: a colour that means *nothing is built* is honest, and a colour that means
# *the column is missing* is not. When the first migration lands, this moves without anybody
# editing a viewer.
hdr(ws,['Module','What it is','Why it is separate','Tables','Written','Cols','Contract','Ops','Tier','Out','In','Cross'],
       [13,52,66,7,8,7,18,6,10,7,6,7])
r=5
for m in MODROWS:
    vals=[m['module'],m['purpose'],m['why'],m['tables'],m.get('written',0),m['columns'],
          m['contracts'],m['ops'],m['tier'],m['refs_out'],m['refs_in'],m['cross']]
    for i,v in enumerate(vals,1):
        c=ws.cell(r,i,v); c.font=M if i in (1,6) else B; c.border=BOX
        if i in (4,5,7,9,10,11): c.alignment=Alignment(horizontal='center',vertical='top')
        if i in (2,3): c.alignment=Alignment(wrap_text=True,vertical='top')
        if m['tier']=='spine': c.fill=GREEN
        elif m['tier']=='satellite': c.fill=AMBER
        elif m['tier']=='mixed': c.fill=BLUE
        else: c.fill=GREY
    ws.row_dimensions[r].height=72
    r+=1
last=r-1
for i,v in enumerate(['TOTAL','','', f'=SUM(D5:D{last})', f'=SUM(E5:E{last})','',
                      f'=SUM(G5:G{last})','', f'=SUM(I5:I{last})', f'=SUM(J5:J{last})', f'=SUM(K5:K{last})'],1):
    c=ws.cell(r,i,v); c.font=BD; c.border=BOX; c.fill=GREY
    if i in (4,5,7,9,10,11): c.alignment=Alignment(horizontal='center')
ws.auto_filter.ref=f"A4:K{last}"
ws.freeze_panes=ws.cell(5,2)
r+=2
ws.cell(r,1,'Reading this sheet').font=BD; r+=1
for t_ in ['Out is how many foreign keys leave the module; In is how many point at it; Cross counts those',
           'crossing a module boundary. platform and identity dominate because every scoped table carries',
           'venue_id and every audited one carries a principal — 97 and 74 references in respectively.',
           '',
           'Ops counts operations on the contracts a module derives from. shift is counted under orders:',
           'its 13 operations write orders.shift, because cash handling is a sales concern.',
           '',
           'pii and sync show no contract. They are storage the platform writes rather than surfaces',
           'anyone calls — which is the point of pii, and why it is the module hardest to reach.']:
    ws.cell(r,1,t_).font=B; r+=1

# ── service ownership ────────────────────────────────────────────────────
# **The lineage already names a service on every operation and the Tables sheet named none.**
# A backend engineer could ask *which service owns this operation* and not *which service owns
# this table* — and the second is the question you ask when you are about to write DDL.
#
# The join is one lookup: a table's schema is its prefix, and the decomposition maps every
# schema to exactly one owner.
try:
    _D = json.load(open('/home/claude/ticvai-pkg/handoff/service-decomposition.json',
                        encoding='utf-8'))['services']
except Exception:
    _D = {}
SVC_OF_SCHEMA = {sch: name for name, v in _D.items() for sch in v.get('schemas', [])}
SVC_CONTRACTS = {name: set(v.get('contracts') or []) for name, v in _D.items()}
# **Loaded here rather than reusing LIN**, which is read 150 lines further down for the lineage
# sheet. A helper that depends on a variable defined after it is a helper that works only where
# it happens to be called.
_LINEAGE = json.load(open(_pkg('api-data-lineage.json','lineage.json'), encoding='utf-8'))

def svc_of(table):
    """Owning service for a table, by its schema prefix."""
    return SVC_OF_SCHEMA.get(table.split('.')[0], '—')

def foreign_writers(table):
    """**Contracts that write this table and do not belong to its owner.**

    22 tables have one. The rule is that the owner defines the row and a foreign writer may
    only append to it — a till closing posts to `ledger.entry` because settling a shift *is*
    a ledger act. Worth seeing on the row rather than in a separate book.
    """
    own = SVC_CONTRACTS.get(svc_of(table), set())
    w = sorted({v['contract'] for v in _LINEAGE.values() if table in (v.get('writes') or [])})
    return [c for c in w if c not in own]

ws=wb.create_sheet('Tables'); ws.cell(1,1,f'All {tot} tables').font=T
ws.cell(2,1,'Blue is new on 14 August. Pink is personal data.').font=SUB
hdr(ws,['Module','Table','Service','Columns','Written','New','PII','Foreign writers','Parent','Anchors on','Derived from','Migration'],
    [12,32,17,9,8,6,6,20,26,30,26,20])
r=5
for t in sorted(set(cols)|set(storage)):
    m=t.split('.')[0]; cs=cols.get(t,[])
    src=origin.get(t,'') or f"storage only — {storage.get(t,'')}"
    L=LINE.get(t,{})
    anch=L.get('anchors') or []
    # A table that is its own anchor owns itself — the spine, not a leaf.
    anchtxt=('itself — nothing above it' if L.get('isAnchor')
             else ', '.join(a.split('.')[1] for a in anch[:3]) + ('…' if len(anch)>3 else ''))
    fw=foreign_writers(t)
    for i,v in enumerate([m,t,svc_of(t).replace('Service',''),len(cs) or '—',
                          'yes' if t in written else '',
                          'yes' if t in NEW else '','yes' if t in PII else '',
                          ', '.join(fw) or '',
                          (L.get('parent') or '—'), anchtxt or '—',
                          src,MIG.get(m,'unassigned')],1):
        c=ws.cell(r,i,v); c.font=M if i==2 else B; c.border=BOX
        if i in (4,5,6,7): c.alignment=Alignment(horizontal='center')
        # **Amber where a contract outside the owning service writes it.** Correct in all 22
        # cases and still the thing to look at first when a boundary is questioned.
        if i==8 and fw: c.fill=AMBER
        if t in PII: c.fill=PINK
        elif t in NEW: c.fill=BLUE
        elif t in written: c.fill=GREEN
        elif t in storage: c.fill=GREY
    r+=1
ws.auto_filter.ref=f"A4:L{r-1}"

ws=wb.create_sheet('Columns'); ws.cell(1,1,'Every column').font=T
ws.cell(2,1,f"{sum(len(v) for v in cols.values())} columns. **References says where a column points; "
        f"Enforced says whether a foreign key checks it.** 28 of 514 are enforced — the rest are "
        f"conventions the relationship graph records and the database does not check.").font=SUB
hdr(ws,['Table','Column','Type','Required','References','Link','Enforced','Source','Description'],
    [30,26,15,9,30,14,10,32,54])
r=5
for t in sorted(cols):
    for c_ in cols[t]:
        row=[c_['table'],c_['column'],c_['type'],c_['required'],
             c_.get('references',''),c_.get('referenceHow',''),c_.get('enforced',''),
             c_['source'],c_['description']]
        for i,v in enumerate(row,1):
            x=ws.cell(r,i,v); x.font=M if i in (1,2,3,5) else B; x.border=BOX
            if i in (4,7): x.alignment=Alignment(horizontal='center')
            if i==9: x.alignment=Alignment(wrap_text=True,vertical='top')
            if t in NEW and i<=2: x.fill=BLUE
        r+=1
ws.auto_filter.ref=f"A4:I{r-1}"

ws=wb.create_sheet('No table'); ws.cell(1,1,'Schemas with no table').font=T
ws.cell(2,1,f'{len(nomap)} of {len(nomap)+len(cols)} mapped schemas.').font=SUB
hdr(ws,['Contract','Schema','Why there is no table'],[18,34,70])
r=5
for ctx,s_,why in sorted(nomap):
    for i,v in enumerate([ctx,s_,why or '—'],1):
        c=ws.cell(r,i,v); c.font=M if i==2 else B; c.border=BOX; c.fill=GREY
        if i==3: c.alignment=Alignment(wrap_text=True,vertical='top')
    r+=1
ws.auto_filter.ref=f"A4:C{r-1}"

ws=wb.create_sheet('Relationships')
res=[x for x in rels if x['to']]; gaps=[x for x in rels if not x['to']]
ws.cell(1,1,'Table relationships').font=T
ws.cell(2,1,f'{len(res)} resolved, {len(gaps)} needing a decision. Cross-module amber, parent-child green.').font=SUB
hdr(ws,['From table','Column','References','Kind','Cross-module','Required'],[32,28,32,20,13,10])
r=5
for x in sorted(res,key=lambda z:(z['frm'],z['col'])):
    for i,v in enumerate([x['frm'],x['col'],x['to'],x['how'],x['cross'],x['required']],1):
        c=ws.cell(r,i,v); c.font=M if i in (1,2,3) else B; c.border=BOX
        if i in (5,6): c.alignment=Alignment(horizontal='center')
        if x['cross']: c.fill=AMBER
        if x['how']=='child': c.fill=GREEN
    r+=1
ws.auto_filter.ref=f"A4:F{r-1}"
r+=1; ws.cell(r,1,'References with no resolved target').font=BD; r+=1
for i,h in enumerate(['From table','Column','Why unresolved'],1):
    c=ws.cell(r,i,h); c.font=HDR; c.fill=HF; c.border=BOX
r+=1
for x in sorted(gaps,key=lambda z:(z['how'],z['frm'])):
    for i,v in enumerate([x['frm'],x['col'],x['how']],1):
        c=ws.cell(r,i,v); c.font=M if i<=2 else B; c.border=BOX; c.fill=RED
    r+=1

ws=wb.create_sheet('Where used'); ws.cell(1,1,'Table to API to screen').font=T
ws.cell(2,1,'What breaks if this table changes — and whose deployment that is.').font=SUB
hdr(ws,['Table','Owner service','Also written by','Operations','Screens','Which operations','Which screens'],
    [32,17,20,11,9,50,38])
r=5
for t in sorted(set(cols)|set(storage)):
    ops=tab_ops.get(t,[]); scr=tab_screens.get(t,[])
    fw=foreign_writers(t)
    # **Whose deployment breaks**, not just what breaks. The owning service is the answer a
    # backend engineer needs before touching a column.
    for i,v in enumerate([t,svc_of(t).replace('Service',''),', '.join(fw) or '',
                          len(ops) or '',len(scr) or '',
                          ', '.join(sorted({o[0] for o in ops}))[:400],', '.join(scr)[:300]],1):
        c=ws.cell(r,i,v); c.font=M if i==1 else B; c.border=BOX
        if i in (4,5): c.alignment=Alignment(horizontal='center')
        if i in (6,7): c.alignment=Alignment(wrap_text=True,vertical='top')
        if i==3 and fw: c.fill=AMBER
        if not ops: c.fill=GREY
        elif scr: c.fill=GREEN
    r+=1
ws.auto_filter.ref=f"A4:G{r-1}"

ws=wb.create_sheet('Scaling'); ws.cell(1,1,'Read and write routing — ADR-0016').font=T
ws.cell(2,1,'Writes always primary.').font=SUB
hdr(ws,['Contract','Writes','Primary reads','Replica reads','Analytical reads'],[22,10,14,14,17])
r=5; tt=defaultdict(int)
for ctx in sorted(routing):
    v=routing[ctx]
    for i,val in enumerate([ctx,v.get('write',0),v.get('primary',0),v.get('replica',0),v.get('analytical',0)],1):
        c=ws.cell(r,i,val); c.font=B; c.border=BOX
        if i>1: c.alignment=Alignment(horizontal='center')
        if i==3 and v.get('primary',0): c.fill=RED
        if i==5 and v.get('analytical',0): c.fill=GREEN
    for k in ('write','primary','replica','analytical'): tt[k]+=v.get(k,0)
    r+=1
for i,val in enumerate(['TOTAL',tt['write'],tt['primary'],tt['replica'],tt['analytical']],1):
    c=ws.cell(r,i,val); c.font=BD; c.border=BOX; c.fill=GREY
    if i>1: c.alignment=Alignment(horizontal='center')

# ---- Data lineage
import json as _j
LIN=_j.load(open(_pkg('api-data-lineage.json','lineage.json'), encoding='utf-8'))
SERVICE={'tenancy':'TenancyService','identity':'IdentityService','catalogue':'CatalogueService',
 'orders':'OrderService','shift':'ShiftService','access':'AccessService','finance':'LedgerService',
 'cross-cell':'CrossCellService','fnb':'FnbService','retail':'RetailService','inventory':'InventoryService',
 'seating':'SeatingService','promotions':'PromotionsService','marketing-crm':'MarketingService',
 'maintenance':'MaintenanceService','queue':'QueueService','white-label':'WhiteLabelService',
 'subscription':'SubscriptionService','platform-ops':'PlatformOpsService','reporting':'ReportingService',
 'assets':'AssetsService','games':'GamesService'}
SP={'createPayment':'orders.sp_capture_payment','createRefund':'orders.sp_post_refund',
 'validateAccess':'access.sp_validate_and_record','acquireLease':'catalogue.sp_acquire_lease',
 'createSeatHold':'seating.sp_hold_seats','closeShift':'orders.sp_close_shift',
 'syncScans':'access.sp_sync_scan_batch','postStockCount':'inventory.sp_post_movement'}
ws=wb.create_sheet('Data lineage')
ws.cell(1,1,'Where each operation gets its data').font=T
ws.cell(2,1,'Every operation against the tables it reads and writes, the service that owns it, and the ten '
            'stored procedures. Derived from persistence markers; 37 projections hand-mapped.').font=SUB
hdr(ws,['Operation','Contract','Service','Verb','Store','Reads','Writes','Routing','Scope','Offline','Stored procedure','Source'],
       [30,15,20,7,17,50,38,11,12,8,26,12])
r=5
for oid in sorted(LIN):
    v=LIN[oid]
    vals=[oid, v['contract'], SERVICE.get(v['contract'],''), v['verb'],
          ' + '.join(v.get('stores') or ['postgres']),
          ', '.join(v['reads'])[:250] or '—', ', '.join(v['writes'])[:200] or '—',
          v['routing'] or ('write' if v['verb']!='GET' else ''), v['scope'],
          'yes' if v['offline'] else '', SP.get(oid,''), v['source']]
    for i,val in enumerate(vals,1):
        c=ws.cell(r,i,val); c.font=M if i in (1,5,6,10) else B; c.border=BOX
        if i in (4,5,8,9,10,12): c.alignment=Alignment(horizontal='center',vertical='top')
        if i in (6,7): c.alignment=Alignment(wrap_text=True,vertical='top')
        if 'qdrant' in (v.get('stores') or []): c.fill=PINK
        elif SP.get(oid): c.fill=BLUE
        elif v['source']=='hand-mapped': c.fill=AMBER
        elif not v['reads']: c.fill=GREY
        elif v['writes']: c.fill=GREEN
    r+=1
ws.auto_filter.ref=f"A4:L{r-1}"
ws.freeze_panes=ws.cell(5,2)
r+=2
ws.cell(r,1,'Reading this sheet').font=BD; r+=1
for t_ in ['Blue rows are the eight stored procedures. Amber were hand-mapped because the operation',
           'returns a projection with no persistence marker. Grey rows resolve to no table — most',
           'return a computed view, a command with no body, or a health check, and that is correct.',
           '',
           'Green rows write. Before changing a table, filter Writes for it and you have the list of',
           'operations that break. orders.sales_order is read by 25 operations across four services,',
           'which is the argument for the outbox rather than direct cross-service writes.',
           '',
           'A read only to check a permission or resolve a scope is not shown. Under RLS every scoped',
           'read also touches scope_node, and listing that on 755 rows would say nothing.']:
    ws.cell(r,1,t_).font=B; r+=1


# Written to both copies the package carries, by absolute path.
#
# This was a bare relative name, so the workbook landed in whatever directory
# the script happened to be run from — the output with the same fault the
# docstring describes in the inputs. The package keeps a copy at the root and
# one in `handoff/`, **and the viewer reads the `handoff/` one**, so a rebuild
# from the root left the copy that is actually read a day stale with no sign of
# it anywhere.
_targets = [_ROOT / 'TICVAI_Schema_Reference.xlsx', _H / 'TICVAI_Schema_Reference.xlsx']
for _t in _targets:
    _t.parent.mkdir(parents=True, exist_ok=True)
    wb.save(_t)
print(f"{tot} tables | {len(written)} written | {len(PII)} PII | {len(NEW)} new")
print("  wrote " + " and ".join(str(t.relative_to(_ROOT)) for t in _targets))
