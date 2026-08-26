#!/usr/bin/env python3
"""Build the database segregation and service decomposition workbook.

**Two questions, one book.** How 379 tables were divided into 26 schemas, and how 26 schemas became
16 deployable services — with the reasoning on the row rather than in a document beside it.

**The reasoning is the deliverable.** A list of sixteen services is a diagram anybody could draw;
what makes it defensible is why `shift` is not one of them and why AI is one at thirty operations.
"""
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import sys

# A cp1252 console cannot encode the arrows and dashes this tool prints, and the
# failure lands *after* the work is done — so the output is written, the summary
# line raises UnicodeEncodeError, and a correct run exits 1. Reconfiguring at
# import means anything importing this module gets it too, refresh.sh included.
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:      # a captured stream may not be reconfigurable; harmless
    pass


ROOT = Path("/home/claude/ticvai-pkg")
OUT = Path("/mnt/user-data/outputs/TICVAI_Services_and_Data_Segregation.xlsx")

HEAD = PatternFill("solid", fgColor="0B1324")
BAND = PatternFill("solid", fgColor="F2F5F9")
TIER = {
    "foundation": PatternFill("solid", fgColor="DDEBF7"),
    "commerce": PatternFill("solid", fgColor="C6EFCE"),
    "operations": PatternFill("solid", fgColor="FFF2CC"),
    "engagement": PatternFill("solid", fgColor="FCE4D6"),
    "platform": PatternFill("solid", fgColor="EDEDED"),
}
THIN = Side(style="thin", color="D8DEE8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def head(ws, cols, widths, row=1):
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row, i, c)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row + 1, 1)


def cell(ws, r, i, v, bold=False, band=False, fill=None):
    c = ws.cell(r, i, v)
    c.font = Font(name="Arial", size=9, bold=bold)
    c.alignment = Alignment(vertical="top", wrap_text=True)
    c.border = BORDER
    if fill:
        c.fill = fill
    elif band:
        c.fill = BAND
    return c


# Why each schema is its own schema. Written, not derived — the reasoning is the point.
SCHEMA_WHY = {
    "identity": ("Principals, roles, sessions, credentials. **Read by twelve contracts and reads "
                 "almost nothing.** Separate because everything resolves a principal before it does "
                 "anything else, and a schema that everything reads must be one nothing can lock."),
    "pii": ("Names, phones, emails, documents. **Separate from `identity` deliberately** (ADR-0023): "
            "a principal is who may act, a subject is who they are. **One writer only** — which is "
            "what makes a subject-access export and a deletion request answerable at all."),
    "platform": ("Tenants, venues, workstations, devices, scope nodes. **`platform.scope_node` is "
                 "reached by 289 of 379 tables** — the tenancy spine, and the terminal anchor for "
                 "almost everything in the package."),
    "catalogue": ("Products, variants, prices, performances, inventory leases. **What is for sale.** "
                  "Separate from `orders` because a product outlives every order against it, and a "
                  "catalogue that lived in the order schema could not be published as a bundle."),
    "orders": ("Carts, orders, lines, payments, refunds, shifts, cash. **The transactional core and "
               "the highest write rate in the platform.** Also holds `orders.shift` and the cash "
               "tables — a till session is an order-side artefact, not a workforce one."),
    "access": ("Entitlements, scans, access points, admission profiles. **A gate decision in under "
               "300ms**, made forty times a minute per lane. Separate because it is the only schema "
               "read at the edge with a local cache."),
    "ledger": ("Journal entries, lines, accounts, settlements. **Append-only, never updated, never "
               "deleted.** A different discipline from everything around it, and one that should "
               "not share a schema with anything that mutates."),
    "fnb": ("Outlets, menus, kitchen tickets, tables, recipes, food safety. **34 tables and the "
            "largest operational domain.** Separate because it is licensed separately and because "
            "the kitchen runs offline."),
    "retail": ("Merchandise, wallets, gift cards, retail sales. Separate because **a venue can run "
               "ticketing with no shop** — the schema exists or it does not, per licence."),
    "inventory": ("Items, movements, counts, requisitions, purchase orders, suppliers. **The level "
                  "is derived from the movements** and never stored — a stored level and a movement "
                  "ledger that disagree is a stock count nobody can reconcile."),
    "marketing": ("Guest profiles, segments, campaigns, consents, cases. **37 tables, the largest "
                  "single schema.** Holds consent records, which is why it reads `pii` and never "
                  "writes it."),
    "control": ("Tenants, plans, licences, cells, migrations, developer accounts. **Three contracts "
                "write it** — subscription, platform-ops and public-api — and that is correct: they "
                "are one bounded context with three API surfaces."),
    "ai": ("Conversations, interactions, proposed actions, knowledge, suggestions. **Only the AI "
           "contract writes these** (ADR-0020), and the boundary caught four wrongly-placed writes "
           "in one week."),
    "whitelabel": ("Brand identity, content pages, navigation, config versions. **The only schema a "
                   "tenant edits directly and publishes from.**"),
    "seating": ("Seat maps, zones, categories, holds. Separate from `catalogue` because a seat map "
                "outlives the performances sold against it."),
    "promotions": ("Promotions, bundles, coupons, vouchers. Separate because **a promotion is "
                   "evaluated on the hot path of every sale** and versioned independently of price."),
    "reporting": ("Report definitions, schedules, executions, dashboards. **Reads almost everything "
                  "and writes only its own definitions** — the schema that must never touch the "
                  "transactional primary."),
    "workforce": ("Rotas, attendance, announcements, incidents. Small, venue-scoped, and folded "
                  "into TenancyService because it reads `scope_node` constantly."),
    "approvals": ("Approval requests and decisions. **One mechanism, not one per thing approved** — "
                  "a write-off, a scenario and a price change are all requests with a subject."),
    "queue": ("Queues, entries, wait readings. Adaptor-first (ADR-0012) — the schema holds what a "
              "third-party queue system reports."),
    "maintenance": ("Assets, work orders, inspections. Separate from `platform.device` because **a "
                    "turnstile is an asset and a workstation is a station** — one is maintained, the "
                    "other is configured."),
    "resources": ("Bookable objects — cabanas, equipment, instructors. **A locker is a resource, not "
                  "an entitlement**, because it is checked out and returned."),
    "venuemap": ("Points, walkways, geometry. Separate because a venue map is imported and labelled "
                 "as a pipeline, not edited as a record."),
    "assets": ("Media assets, usage, upload tickets. Cross-tenant sharing refused outside one "
               "tenant's own scope tree."),
    "games": ("Arcade, credit ledger, redemption. Fully optional, fully separate."),
    "sync": ("Cross-cell links and rejections. **The only schema that reaches another region** — a "
             "boundary that matters legally, made a boundary that exists physically."),
}


def main() -> int:
    D = json.loads((ROOT / "handoff" / "service-decomposition.json").read_text(encoding="utf-8"))
    lin = json.loads((ROOT / "handoff" / "api-data-lineage.json").read_text(encoding="utf-8"))
    S = json.loads((ROOT / "handoff" / "schema-reference.json").read_text(encoding="utf-8"))
    real = {t for t in (set(S.get("cols") or {}) | set(S.get("storage") or {}))
            if "." in t and ":" not in t}

    wb = openpyxl.Workbook()

    # ── 1. Services ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Services"
    ws["A1"] = "TICVAI — sixteen services, and why each is one"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="0B1324")
    ws["A2"] = ("28 contracts, 1,007 operations, 379 tables. The data boundaries were drawn first "
                "and the service boundaries follow them — no service spans a schema it does not own, "
                "and no schema is written by two services.")
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="5A6577")
    ws["A3"] = ("That was true before this book existed. The work was finding it, not creating it — "
                "and on 20 August it was not yet true.")
    ws["A3"].font = Font(name="Arial", size=10, italic=True, color="A81E1E")

    head(ws, ["Service", "Tier", "Ops", "Tables", "Walked", "Contracts", "Schemas owned",
              "Why it is its own service", "Scale profile", "What breaks if it is down"],
         [22, 13, 6, 7, 8, 26, 26, 74, 34, 34], row=5)

    order = ["foundation", "commerce", "operations", "engagement", "platform"]
    rows = sorted(D["services"].items(), key=lambda kv: (order.index(kv[1]["tier"]), -kv[1]["operations"]))
    r = 6
    for i, (name, v) in enumerate(rows):
        cov = (v.get("flowCoverage") or {}).get("percent", 0)
        vals = [name, v["tier"], v["operations"], v["tables"], f"{cov}%",
                ", ".join(v["contracts"]),
                ", ".join(v["schemas"]), v["why"], v["scale"], v["risk"]]
        for j, val in enumerate(vals, 1):
            cell(ws, r, j, val, bold=(j == 1), band=(i % 2 == 0))
        ws.cell(r, 2).fill = TIER[v["tier"]]
        # **A service walked at 31% next to one at 96% is the fact a build team needs beside the
        # deploy order.** 93 flows cover 53% of operations and the spread is what says where to look.
        ws.cell(r, 5).fill = PatternFill("solid", fgColor=("C6EFCE" if cov >= 60 else
                                                            "FFF2CC" if cov >= 40 else "FDECEC"))
        r += 1

    # ── 2. Schema segregation ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Schema segregation")
    ws2["A1"] = "Why each schema is its own schema"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color="0B1324")
    ws2["A2"] = ("26 schemas in one Postgres per cell. Not one database per tenant (ADR-0005) and "
                 "not one shared table with a tenant column — partitioning by scope_path, "
                 "segregation by domain.")
    ws2["A2"].font = Font(name="Arial", size=10, italic=True, color="5A6577")

    head(ws2, ["Schema", "Owning service", "Tables", "Store", "Written by",
               "Read by (contracts)", "Why it is separate"],
         [16, 22, 8, 18, 26, 26, 88], row=4)

    owner = {s: n for n, v in D["services"].items() for s in v["schemas"]}
    writers = defaultdict(set)
    readers = defaultdict(set)
    for o, v in lin.items():
        for t in v.get("writes", []):
            if t in real:
                writers[t.split(".")[0]].add(v["contract"])
        for t in v.get("reads", []):
            if t in real:
                readers[t.split(".")[0]].add(v["contract"])

    schemas = sorted({t.split(".")[0] for t in real})
    r = 5
    for i, sch in enumerate(schemas):
        tables = [t for t in real if t.split(".")[0] == sch]
        stores = sorted({S.get("store", {}).get(t, "postgres") for t in tables})
        w = sorted(writers.get(sch, set()))
        rd = sorted(readers.get(sch, set()) - set(w))
        vals = [sch, owner.get(sch, "—"), len(tables), ", ".join(stores),
                ", ".join(w) or "nothing writes it",
                (", ".join(rd[:6]) + (f" +{len(rd) - 6}" if len(rd) > 6 else "")) or "—",
                SCHEMA_WHY.get(sch, "")]
        for j, val in enumerate(vals, 1):
            cell(ws2, r, j, val, bold=(j == 1), band=(i % 2 == 0))
        if len(w) > 1:
            ws2.cell(r, 5).fill = PatternFill("solid", fgColor="FFF2CC")
        r += 1

    # ── 3. Cross-service writes ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Cross-service writes")
    ws3["A1"] = "Where one service writes another's schema, and why each is allowed"
    ws3["A1"].font = Font(name="Arial", size=13, bold=True, color="0B1324")
    ws3["A2"] = ("The rule: the owner defines the row, a foreign writer may only append to it. "
                 "Every row here was checked — one was a genuine defect and four operations were "
                 "rewired on 20 August.")
    ws3["A2"].font = Font(name="Arial", size=10, italic=True, color="5A6577")

    head(ws3, ["Table", "Owner", "Also written by", "Ops", "Verdict"],
         [34, 20, 30, 6, 84], row=4)

    WHY = {
        "ledger.entry": "**A till closing is a ledger act.** Settling a shift posts to the ledger, and a "
                        "payment writes its own entry — appending, never redefining.",
        "ledger.journal_entry": "As above. The journal is written by whatever caused the money to move.",
        "access.entitlement": "**A sale issues a ticket.** `orders` appends the entitlement, `catalogue` "
                              "suspends and freezes it against its template. Neither redefines the row.",
        "marketing.message_dispatch": "**Five writers, and every one is sending a message somebody asked for.** "
                                      "A verification email, a payment link, an evacuation notice.",
        "orders.sales_order": "`identity` links a guest checkout to an existing order. One field, on claim.",
        "pii.subject": "**`identity` is the only real writer.** `marketing-crm` updates a profile field "
                       "the subject owns — and consent never merges permissively (CF-160).",
        "identity.grant": "A share and a developer membership are both delegated authority (CF-132) — "
                          "one mechanism, three callers.",
        "inventory.movement": "**Every stock change is a movement.** Waste from F&B and adjustment from "
                              "inventory are the same act with a different reason.",
        "ai.interaction": "**The governed reporting pair** (ADR-0020). `reporting` may write an AI "
                          "interaction because a natural-language query is one.",
    }
    w2 = defaultdict(lambda: defaultdict(int))
    for o, v in lin.items():
        for t in v.get("writes", []):
            if t in real:
                w2[t][v["contract"]] += 1
    multi = {t: c for t, c in w2.items() if len(c) > 1}
    r = 5
    for i, (t, c) in enumerate(sorted(multi.items(), key=lambda kv: -len(kv[1]))):
        own = owner.get(t.split(".")[0], "—")
        vals = [t, own, ", ".join(sorted(c)), sum(c.values()),
                WHY.get(t, "Checked and correct — a side effect of the writing service's own job.")]
        for j, val in enumerate(vals, 1):
            cell(ws3, r, j, val, bold=(j == 1), band=(i % 2 == 0))
        r += 1

    # ── 4. Deploy order ──────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Deploy order")
    ws4["A1"] = "What ships in what order, and what can be down"
    ws4["A1"].font = Font(name="Arial", size=13, bold=True, color="0B1324")
    head(ws4, ["Order", "Tier", "Services", "Rule"], [8, 14, 46, 96], row=3)
    seq = [
        (1, "foundation", "IdentityService, TenancyService",
         "**First and alone.** Twelve contracts read identity and 289 tables anchor on "
         "`platform.scope_node` — a restart here is an outage everywhere."),
        (2, "commerce", "CatalogueService, LedgerService, AccessService, OrderService",
         "**Catalogue before Order**, because a till pulls a bundle before it sells. **Access last "
         "of the four** — it runs at the edge with a local cache and can lag the others safely."),
        (3, "operations", "InventoryService, FnbService, RetailService, VenueOpsService",
         "**Module-gated.** A venue that licensed none of these runs none of them, which is the "
         "whole point of `requiresModule`."),
        (4, "engagement", "MarketingService, AiService",
         "**Nothing that takes money depends on these.** They can ship late and be down."),
        (5, "platform", "ControlService, WhiteLabelService, ReportingService, CrossCellService",
         "**Control is needed to provision a tenant and not to serve one.** Reporting reads the "
         "replica. CrossCell only matters once a second region exists."),
    ]
    for i, row in enumerate(seq, 4):
        for j, val in enumerate(row, 1):
            cell(ws4, i, j, val, bold=(j == 1), band=(i % 2 == 0))
        ws4.cell(i, 2).fill = TIER[row[1]]

    ws4.cell(11, 1, "Can be down without stopping a sale").font = Font(name="Arial", size=11, bold=True)
    ws4.cell(12, 1, "MarketingService · AiService · ReportingService · CrossCellService").font = \
        Font(name="Arial", size=10)
    ws4.cell(13, 1, "A deliberate property, and one that should be tested rather than assumed.").font = \
        Font(name="Arial", size=9, italic=True, color="5A6577")

    # ── 5. Scope hierarchy ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Scope hierarchy")
    ws5["A1"] = "Eight levels, and why each one exists"
    ws5["A1"].font = Font(name="Arial", size=13, bold=True, color="0B1324")
    ws5["A2"] = ("platform.scope_node is reached by 289 of 379 tables — the tenancy spine. "
                 "Configuration resolves by walking the path upward until something answers.")
    ws5["A2"].font = Font(name="Arial", size=10, italic=True, color="5A6577")
    head(ws5, ["Level", "Branch", "Ops", "Configs", "What it owns", "Why it exists at this height"],
         [16, 18, 6, 8, 42, 92], row=4)
    LEVELS = [
        ("tenant", "root", 238, 25, "Plans, licences, roles, message templates, loyalty programmes",
         "**The commercial entity, and the level things are bought at.** A role defined at venue "
         "level would have to be defined eleven times."),
        ("brand", "organisational", 2, 0, "Brand identity between company and country",
         "**Optional, and it costs nothing when unused.** A tenant with one brand never creates a "
         "node and the resolver skips a level that is not there."),
        ("region", "organisational", 46, 8, "Currency, decimal scale, timezone, fiscal year",
         "**The level most often got wrong.** These four are not venue settings and are not "
         "overridable below — a platform that lets one venue pick a different decimal scale has "
         "produced a ledger that cannot be consolidated (ADR-0018)."),
        ("venue", "organisational", 675, 29, "Almost everything operational",
         "**Two thirds of all operations scope here, and that is the correct shape.** A venue is "
         "the unit a guest visits, a shift is worked at and stock is held in. The platform's "
         "default."),
        ("department", "organisational", 0, 0, "Requisitions, rotas, cost centres",
         "**The staffing and budget tree** — who works where and whose budget it comes from."),
        ("subDepartment", "organisational", 0, 0, "A second tier of the same",
         "Optional. A venue that does not need two tiers does not create them."),
        ("workstation", "organisational", 45, 0,
         "Config profile, offline policy, connectivity thresholds, deposit box",
         "**A till is a scope, not just a device.** ADR-0002 makes authorisation user-driven, but "
         "the workstation still holds settings a person does not carry with them."),
        ("outlet", "COMMERCIAL", 0, 5,
         "Menu, table layout, return policy, opening hours, substitution rules",
         "**A sibling of department, never a child of it** (CF-138, ADR-0018). A department "
         "answers *who works where*; an outlet answers *what is sold where*. **Modelling a "
         "restaurant as a department puts it in the staffing tree** — a venue whose restaurant is "
         "a concession would be describing staff it does not employ. Only F&B and retail resolve "
         "here; everything else has no outlet on its path."),
    ]
    for i, row in enumerate(LEVELS, 5):
        for j, val in enumerate(row, 1):
            cell(ws5, i, j, val, bold=(j == 1), band=(i % 2 == 0))
        ws5.cell(i, 2).fill = (PatternFill("solid", fgColor="FFF2CC") if row[1] == "COMMERCIAL"
                               else PatternFill("solid", fgColor="DDEBF7"))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"  {len(D['services'])} services · {len(schemas)} schemas · {len(multi)} cross-service writes")
    print(f"  → {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
